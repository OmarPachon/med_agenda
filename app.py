#!/usr/bin/env python3
"""
Agendamiento-Med Elite v2.0 - Acceso por Profesional en Red Local
Sistema de Gestión de Citas Médicas (CSV/TXT)
© 2025 Omar Alberto Pachon Pereira
Todos los derechos reservados.
"""
import os
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from functools import wraps
from werkzeug.security import check_password_hash, generate_password_hash
import csv
import sys
csv.field_size_limit(sys.maxsize)
from datetime import datetime, date, timedelta
from collections import defaultdict
from io import BytesIO
import pandas as pd
from fpdf import FPDF
from fpdf.enums import XPos, YPos
app = Flask(__name__)
app.secret_key = 'cambia_esta_clave_secreta_por_una_muy_segura_en_produccion_12345!'
# === PREVENIR CACHÉ EN TODAS LAS RESPUESTAS ===
@app.after_request
def add_no_cache_headers(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, private, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    response.headers['Surrogate-Control'] = 'no-store'
    response.headers['Vary'] = '*'
    return response
# === Configuración de usuarios ===
USUARIOS_ADMIN = {
    "admin": generate_password_hash("1234"),
    "usuario1": generate_password_hash("Usuario1_2025"),
}
def login_requerido(f):
    @wraps(f)
    def decorada(*args, **kwargs):
        if 'usuario' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorada
# === RUTA DE LOGIN ===
@app.route("/login", methods=["GET", "POST"])
def login():
    profesionales = []
    if os.path.exists("data/Profesionales.txt"):
        with open("data/Profesionales.txt", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines[1:]:
                partes = line.strip().split("|")
                if partes and partes[0].strip() and partes[0].strip() not in USUARIOS_ADMIN:
                    profesionales.append(partes[0].strip())
    if request.method == "POST":
        nombre_usuario = request.form.get("usuario", "").strip()
        contrasena = request.form.get("contrasena", "")
        if nombre_usuario in USUARIOS_ADMIN:
            if check_password_hash(USUARIOS_ADMIN[nombre_usuario], contrasena):
                session['usuario'] = nombre_usuario
                session['rol'] = 'admin'
                return redirect(url_for("index"))
            else:
                return render_template("login.html", error="❌ Contraseña incorrecta para administrador", profesionales=profesionales)
        if nombre_usuario in profesionales:
            session['usuario'] = nombre_usuario
            session['rol'] = 'profesional'
            return redirect(url_for("index"))
        return render_template("login.html", error="❌ Usuario no encontrado en el sistema", profesionales=profesionales)
    return render_template("login.html", profesionales=profesionales)
@app.route("/logout")
def logout():
    session.pop('usuario', None)
    session.pop('rol', None)
    response = redirect(url_for('login'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response
# === Configuración de campos y constantes ===
CAMPOS_AGENDA = [
    "Tipo_Doc", "Num_Doc", "Nombre_Completo", "Edad", "Genero",
    "REGIMEN", "TIPO_AFILIADO", "CODIGO_EPS", "DEPARTAMENTO", "CIUDAD", "ZONA", "ESTADO",
    "Celular", "Email", "Dx_Codigo", "Dx_Descripcion",
    "CUPS", "Servicio", "Profesional", "Fecha", "Hora",
    "Cantidad_Total", "Frecuencia_Semanal", "Duracion_Meses", "Observacion",
    "Fecha_Registro", "Estado"
]
FESTIVOS_2026 = {
    "2026-01-01", "2026-03-23", "2026-04-02", "2026-04-03",
    "2026-01-12", "2026-06-15", "2026-07-20", "2026-08-07",
    "2026-08-17", "2026-10-12", "2026-11-02", "2026-11-16",
    "2026-12-08", "2026-12-25", "2026-05-01", "2026-05-18", "2026-06-08", "2026-06-29"
}
TERAPIAS_REPETITIVAS = {
    "933901", "937000", "937101", "937203", "937400",
    "938303", "944301", "938610", "938611", "938612"
}
def generar_bloques(hora_inicio, hora_fin):
    bloques = []
    h_inicio = datetime.strptime(hora_inicio, "%H:%M")
    h_fin = datetime.strptime(hora_fin, "%H:%M")
    while h_inicio <= h_fin:
        bloques.append(h_inicio.strftime("%H:%M"))
        h_inicio += timedelta(minutes=30)
    return bloques
BLOQUES_MANANA = generar_bloques("07:00", "11:30")
BLOQUES_TARDE = generar_bloques("13:00", "16:30")
TODOS_BLOQUES = BLOQUES_MANANA + BLOQUES_TARDE
def es_dia_habil(fecha_str):
    try:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        if fecha.weekday() >= 5:
            return False
        if fecha_str in FESTIVOS_2026:
            return False
        return True
    except:
        return False
# === FUNCIONES DE CARGA DE DATOS ===
def cargar_pacientes(filepath):
    pacientes = {}
    if not os.path.exists(filepath):
        return pacientes
    with open(filepath, encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="|")
        for row in reader:
            if row.get("Estado") == "AC":
                clave = f"{row['Tipo_Documento']}|{row['Numero_Documento']}"
                fecha_nac = row.get("Fecha_Nacimiento", "").strip()
                try:
                    edad = (date.today() - datetime.strptime(fecha_nac, "%Y/%m/%d").date()).days // 365
                except:
                    edad = None
                pacientes[clave] = {
                    "Primer_Nombre": row.get("Primer_Nombre", "").strip(),
                    "Segundo_Nombre": row.get("Segundo_Nombre", "").strip(),
                    "Primer_Apellido": row.get("Primer_Apellido", "").strip(),
                    "Segundo_Apellido": row.get("Segundo_Apellido", "").strip(),
                    "Fecha_Nacimiento": fecha_nac,
                    "Edad": edad,
                    "Genero": row.get("Genero", "").strip(),
                    "REGIMEN": row.get("REGIMEN", "").strip(),
                    "TIPO_AFILIADO": row.get("TIPO_AFILIADO", "").strip(),
                    "CODIGO_EPS": row.get("Codigo_EPS", "").strip(),
                    "DEPARTAMENTO": row.get("Departamento", "").strip(),
                    "CIUDAD": row.get("Ciudad", "").strip(),
                    "ZONA": row.get("Zona", "").strip(),
                    "ESTADO": row.get("Estado", "").strip(),
                }
    return pacientes
def cargar_dx(filepath):
    dx = {}
    if not os.path.exists(filepath):
        return dx
    with open(filepath, encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="|")
        for row in reader:
            dx[row["Dx"].strip()] = row["DIAGNOSTICO"].strip()
    return dx
def cargar_servicios(filepath):
    servicios = {}
    if not os.path.exists(filepath):
        return servicios
    with open(filepath, encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="|")
        for row in reader:
            servicios[row["CUPS_ISS"].strip()] = row["SERVICIO"].strip()
    return servicios
def cargar_profesionales(filepath):
    profesionales = []
    if not os.path.exists(filepath):
        return profesionales
    with open(filepath, encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[1:]:
            partes = line.strip().split("|")
            if partes and partes[0].strip():
                profesionales.append(partes[0].strip())
    return sorted(profesionales)
# ✅ FUNCIÓN CLAVE: Cargar agenda FILTRADA (MODIFICADA - Regla Deglución = Individual como Autismo)
def cargar_agenda_desde_csv_filtrada(profesional_filtro=None, solo_activas=True):
    agenda = {}
    if not os.path.exists("data/agenda.csv"):
        return agenda
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            estado = row.get("Estado", "activa").strip().lower()
            if solo_activas and estado != "activa":
                continue
            profesional = row.get("Profesional", "").strip()
            fecha = row.get("Fecha", "").strip()
            hora = row.get("Hora", "").strip()
            dx = row.get("Dx_Codigo", "").strip().upper()
            cups = row.get("CUPS", "").strip()  # ✅ LEER CUPS
            if profesional_filtro and profesional != profesional_filtro:
                continue
            if profesional and fecha and hora:
                clave = f"{profesional}|{fecha}|{hora}"
                if clave not in agenda:
                    agenda[clave] = []
                es_autismo = dx == "F840"
                es_deglucion = cups == "937203"  # ✅ DEGLUCIÓN = INDIVIDUAL (como autismo)
                row["_es_autismo"] = es_autismo
                row["_es_deglucion"] = es_deglucion
                agenda[clave].append(row)
    return agenda
def contar_sesiones_realizadas(tipo_doc, num_doc, cups):
    if not os.path.exists("data/agenda.csv"):
        return 0
    contador = 0
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if (row.get("Tipo_Doc") == tipo_doc and
                row.get("Num_Doc") == num_doc and
                row.get("CUPS") == cups and
                row.get("Estado", "activa") == "activa"):
                contador += 1
    return contador
def obtener_celular_profesional(nombre_profesional):
    celular = "3"
    if not os.path.exists("data/Profesionales.txt"):
        return celular
    with open("data/Profesionales.txt", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[1:]:
            partes = [p.strip() for p in line.split("|")]
            if len(partes) >= 4 and partes[0] == nombre_profesional:
                celular = partes[3]
                break
    return celular
def cargar_celulares_profesionales():
    celulares = {}
    if not os.path.exists("data/Profesionales.txt"):
        return celulares
    with open("data/Profesionales.txt", encoding="utf-8") as f:
        lines = f.readlines()
        for line in lines[1:]:
            partes = [p.strip() for p in line.split("|")]
            if len(partes) >= 4:
                nombre = partes[0]
                cel = partes[3]
                cel = "".join(filter(str.isdigit, cel))
                if cel and not cel.startswith("57"):
                    cel = "57" + cel
                celulares[nombre] = cel
    return celulares
def eliminar_citas_bloque_existente(tipo_doc, num_doc, cups, profesional):
    if not os.path.exists("data/agenda.csv"):
        return 0
    registros_actualizados = []
    eliminadas = 0
    fieldnames = None
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or CAMPOS_AGENDA
        for row in reader:
            if not row or not any(row.values()):
                continue
            estado = row.get("Estado", "activa")
            if estado not in ("activa", "cancelada", "no_asistio"):
                estado = "activa"
            if (row.get("Tipo_Doc") == tipo_doc and
                row.get("Num_Doc") == num_doc and
                row.get("CUPS") == cups and
                row.get("Profesional") == profesional and
                estado == "activa"):
                eliminadas += 1
            else:
                fila_completa = {campo: row.get(campo, "") for campo in CAMPOS_AGENDA}
                registros_actualizados.append(fila_completa)
    if registros_actualizados or not os.path.exists("data/agenda.csv"):
        with open("data/agenda.csv", "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=CAMPOS_AGENDA)
            writer.writeheader()
            writer.writerows(registros_actualizados)
    else:
        with open("data/agenda.csv", "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=CAMPOS_AGENDA)
            writer.writeheader()
    return eliminadas
def generar_fechas_terapia(fecha_inicio, frecuencia_semanal, cantidad_total):
    if cantidad_total <= 0:
        return []
    fecha_inicio_date = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
    while not es_dia_habil(fecha_inicio_date.isoformat()):
        fecha_inicio_date += timedelta(days=1)
    fechas = [fecha_inicio_date.isoformat()]
    if cantidad_total == 1:
        return fechas
    if frecuencia_semanal == 1:
        dia_semana_inicio = fecha_inicio_date.weekday()
        semana = 1
        while len(fechas) < cantidad_total:
            propuesta = fecha_inicio_date + timedelta(weeks=semana)
            if propuesta.weekday() == dia_semana_inicio and es_dia_habil(propuesta.isoformat()):
                fechas.append(propuesta.isoformat())
            semana += 1
            if semana > 100:
                break
        return fechas[:cantidad_total]
    if frecuencia_semanal == 2:
        dia_inicio = fecha_inicio_date.weekday()
        if dia_inicio == 0:
            patron = [0, 2]
        elif dia_inicio == 1:
            patron = [1, 3]
        elif dia_inicio == 2:
            patron = [2, 4]
        elif dia_inicio in (3, 4):
            patron = [3, 4]
        else:
            patron = [0, 2]
        lunes_base = fecha_inicio_date - timedelta(days=dia_inicio)
        semana_offset = 0
        while len(fechas) < cantidad_total:
            for wd in patron:
                propuesta = lunes_base + timedelta(weeks=semana_offset) + timedelta(days=wd)
                if propuesta >= fecha_inicio_date and es_dia_habil(propuesta.isoformat()):
                    if propuesta.isoformat() not in fechas:
                        fechas.append(propuesta.isoformat())
                    if len(fechas) >= cantidad_total:
                        break
            semana_offset += 1
            if semana_offset > 100:
                break
        return fechas[:cantidad_total]
    if frecuencia_semanal == 3:
        dias_semana = [0, 2, 4]
    elif frecuencia_semanal == 4:
        dias_semana = [0, 1, 3, 4]
    else:
        dias_semana = [0, 1, 2, 3, 4]
    semana = 0
    while len(fechas) < cantidad_total:
        for dia in dias_semana:
            propuesta = fecha_inicio_date + timedelta(weeks=semana) + timedelta(days=dia)
            if propuesta >= fecha_inicio_date and es_dia_habil(propuesta.isoformat()):
                if propuesta.isoformat() not in fechas:
                    fechas.append(propuesta.isoformat())
                if len(fechas) >= cantidad_total:
                    break
        semana += 1
        if semana > 100:
            break
    return fechas[:cantidad_total]
os.makedirs("data", exist_ok=True)
os.makedirs("templates", exist_ok=True)
os.makedirs("static/css", exist_ok=True)
os.makedirs("static/js", exist_ok=True)
os.makedirs("static/img", exist_ok=True)

def _iniciar_csv_gas():
    """Crea los CSV de GAS si no existen (con estructura actualizada)"""
    if not os.path.exists("data/gas_metas.csv"):
        with open("data/gas_metas.csv", "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter="|")
            writer.writerow([
                "Meta_ID","Paciente_TipoDoc","Paciente_NumDoc","Descripcion",
                "Dominio","Peso","Fecha_Creacion","Fecha_Evaluacion","Estado",
                "Profesional_Creacion","Linea_Base_Puntuacion"  # ✅ NUEVA COLUMNA
            ])
        print("✅ gas_metas.csv creado")
    
    if not os.path.exists("data/gas_niveles.csv"):
        with open("data/gas_niveles.csv", "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter="|")
            writer.writerow(["Meta_ID","Puntuacion","Criterio_Observable","Metodo_Verificacion"])
        print("✅ gas_niveles.csv creado")
    
    if not os.path.exists("data/gas_evaluaciones.csv"):
        with open("data/gas_evaluaciones.csv", "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter="|")
            writer.writerow([
                "Eval_ID","Meta_ID","Fecha_Eval","Puntuacion_Lograda",
                "Evidencia","Profesional_Evaluador","Tipo_Eval"  # ✅ NUEVA COLUMNA
            ])
        print("✅ gas_evaluaciones.csv creado")
import math

def calcular_tscore_gas_multiple(metas_evaluadas, rho=0.30):
    """
    Calcula T-score GAS para múltiples metas evaluadas (Kiresuk & Sherman, 1968)
    
    Args:
        metas_evaluadas: Lista de dicts [{'peso': 2, 'puntuacion': 0}, ...]
        rho: Correlación esperada entre metas (default: 0.30)
    
    Returns:
        dict con t_score, interpretación y detalles del cálculo
    """
    if not metas_evaluadas:
        return {'t_score': None, 'interpretacion': 'Sin metas evaluadas', 'detalles': {}}
    
    # Calcular sumatorias
    sum_w = sum(m['peso'] for m in metas_evaluadas)
    sum_wx = sum(m['peso'] * m['puntuacion'] for m in metas_evaluadas)
    sum_w2 = sum(m['peso']**2 for m in metas_evaluadas)
    
    # Calcular denominador
    denominador = math.sqrt(sum_w2 * (1 - rho) + rho * (sum_w**2))
    
    # Evitar división por cero
    if denominador == 0:
        return {'t_score': 50.0, 'interpretacion': 'Meta lograda (caso especial)', 'detalles': {}}
    
    # Calcular T-score
    t_score = 50 + (sum_wx * 10) / denominador
    
    # Interpretación clínica
    if t_score < 40:
        interpretacion = "Progreso insuficiente 🔴"
    elif t_score < 50:
        interpretacion = "Progreso parcial 🟡"
    elif t_score < 60:
        interpretacion = "Meta lograda ✅ 🟢"
    elif t_score < 70:
        interpretacion = "Superación moderada 🔵"
    else:
        interpretacion = "Resultado excepcional 🟣"
    
    return {
        't_score': round(t_score, 2),
        'interpretacion': interpretacion,
        'detalles': {
            'sum_w': sum_w,
            'sum_wx': sum_wx,
            'sum_w2': sum_w2,
            'denominador': round(denominador, 3),
            'num_metas': len(metas_evaluadas),
            'rho': rho
        }
    }
# ✅ LLAMAR LA FUNCIÓN AL INICIO
_iniciar_csv_gas()

class ReportePDF(FPDF):
    def header(self):
        logo_path = os.path.join("static", "img", "logo.jpg")
        if os.path.exists(logo_path):
            self.image(logo_path, 15, 12, 30)
        self.set_font("Arial", "B", 16)
        self.cell(0, 10, "NEURODESARROLLO", 0, 1, "C")
        self.set_font("Arial", "", 10)
        self.cell(0, 5, "Cra 36 #54-50, Barrio Cabecera, Bucaramanga", 0, 1, "C")
        self.ln(12)
    def footer(self):
        self.set_y(-20)
        self.set_font("Arial", "I", 8)
        self.cell(0, 5, "¡Gracias por confiar en NEURODESARROLLO!", 0, 1, "C")
        self.ln(3)
        self.cell(0, 5, "Sistema Agendamiento-Med Elite v2.0", 0, 1, "C")
# ✅ RUTA PRINCIPAL
@app.route("/")
@login_requerido
def index():
    hoy = date.today().isoformat()
    rol = session.get('rol', 'admin')
    profesional_logueado = session.get('usuario', '')
    if rol == 'profesional':
        profesionales = [profesional_logueado]
    else:
        profesionales = cargar_profesionales("data/Profesionales.txt")
    return render_template("index.html",
        profesionales=profesionales,
        fecha_min=hoy,
        profesional_logueado=profesional_logueado,
        es_profesional=(rol == 'profesional'))
# ✅ RUTA: Agenda general (MODIFICADA - Deglución = Individual como Autismo)
@app.route("/agenda-general", methods=["POST"])
@login_requerido
def agenda_general():
    data = request.get_json()
    fecha = data.get("fecha")
    if not fecha or not es_dia_habil(fecha):
        return jsonify({"error": "Fecha inválida o no hábil"}), 400
    rol = session.get('rol', 'admin')
    profesional_logueado = session.get('usuario', '')
    if rol == 'profesional':
        AGENDA_ACTUAL = cargar_agenda_desde_csv_filtrada(profesional_logueado)
        profesionales_mostrar = [profesional_logueado]
    else:
        AGENDA_ACTUAL = cargar_agenda_desde_csv_filtrada()
        profesionales_mostrar = cargar_profesionales("data/Profesionales.txt")
    resultado = {}
    for profesional in profesionales_mostrar:
        bloques = []
        for hora in TODOS_BLOQUES:
            clave = f"{profesional}|{fecha}|{hora}"
            citas = AGENDA_ACTUAL.get(clave, [])
            ocupados = len(citas)
            tiene_autismo = any(cita.get("_es_autismo") for cita in citas)
            tiene_deglucion = any(cita.get("_es_deglucion") for cita in citas)  # ✅ AGREGADO
            if tiene_autismo:
                color = "morado"
            elif tiene_deglucion:  # ✅ AGREGADO
                color = "amarillo"
            elif ocupados == 0:
                color = "verde"
            elif ocupados == 1:
                color = "azul"
            else:
                color = "rojo"
            bloques.append({"hora": hora, "ocupados": ocupados, "color": color})
        resultado[profesional] = bloques
    return jsonify({"fecha": fecha, "agenda": resultado})
# ✅ RUTA: Estado de agenda (MODIFICADA - Deglución = Individual como Autismo)
@app.route("/estado-agenda", methods=["POST"])
@login_requerido
def estado_agenda():
    data = request.get_json()
    profesional = data.get("profesional")
    fecha = data.get("fecha")
    if not profesional or not fecha:
        return jsonify({"error": "Faltan datos"}), 400
    if not es_dia_habil(fecha):
        return jsonify({"error": "Día no hábil"}), 400
    rol = session.get('rol', 'admin')
    profesional_logueado = session.get('usuario', '')
    if rol == 'profesional' and profesional != profesional_logueado:
        return jsonify({"error": "Acceso denegado. Solo puedes ver tu propia agenda."}), 403
    AGENDA_ACTUAL = cargar_agenda_desde_csv_filtrada(profesional if rol == 'profesional' else None)
    estado = []
    for hora in TODOS_BLOQUES:
        clave = f"{profesional}|{fecha}|{hora}"
        citas = AGENDA_ACTUAL.get(clave, [])
        ocupados = len(citas)
        tiene_autismo = any(cita.get("_es_autismo") for cita in citas)
        tiene_deglucion = any(cita.get("_es_deglucion") for cita in citas)  # ✅ AGREGADO
        if ocupados == 0:
            color = "verde"
        elif tiene_autismo:  # ✅ CAMBIADO: autismo primero
            color = "morado"
        elif tiene_deglucion:  # ✅ AGREGADO
            color = "amarillo"
        elif ocupados == 1:
            color = "azul"
        else:
            color = "rojo"
        estado.append({"hora": hora, "ocupados": ocupados, "color": color, "tiene_autismo": tiene_autismo, "tiene_deglucion": tiene_deglucion})  # ✅ AGREGADO tiene_deglucion
    return jsonify({"estado": estado})
# ✅ RUTA: Agenda filtrada (MODIFICADA - Deglución = Individual como Autismo)
@app.route("/agenda-filtrada", methods=["POST"])
@login_requerido
def agenda_filtrada():
    data = request.get_json()
    fecha = data.get("fecha")
    profesional_filtro = data.get("profesional", "").strip()
    if not fecha or not es_dia_habil(fecha):
        return jsonify({"error": "Fecha inválida o no hábil"}), 400
    rol = session.get('rol', 'admin')
    profesional_logueado = session.get('usuario', '')
    if rol == 'profesional':
        profesional_filtro = profesional_logueado
    AGENDA_ACTUAL = cargar_agenda_desde_csv_filtrada(profesional_filtro if profesional_filtro else None)
    if profesional_filtro:
        profesionales_a_mostrar = [profesional_filtro]
    else:
        profesionales_a_mostrar = [profesional_logueado] if rol == 'profesional' else cargar_profesionales("data/Profesionales.txt")
    resultado = {}
    for profesional in profesionales_a_mostrar:
        bloques = []
        for hora in TODOS_BLOQUES:
            clave = f"{profesional}|{fecha}|{hora}"
            citas = AGENDA_ACTUAL.get(clave, [])
            ocupados = len(citas)
            tiene_autismo = any(cita.get("_es_autismo") for cita in citas)
            tiene_deglucion = any(cita.get("_es_deglucion") for cita in citas)  # ✅ AGREGADO
            if tiene_autismo:
                color = "morado"
            elif tiene_deglucion:  # ✅ AGREGADO
                color = "amarillo"
            elif ocupados == 0:
                color = "verde"
            elif ocupados == 1:
                color = "azul"
            else:
                color = "rojo"
            bloques.append({"hora": hora, "ocupados": ocupados, "color": color})
        resultado[profesional] = bloques
    return jsonify({"fecha": fecha, "agenda": resultado})
@app.route("/validar-paciente", methods=["POST"])
@login_requerido
def validar_paciente():
    data = request.get_json()
    tipo_doc = data.get("tipo_doc", "").strip().upper()
    num_doc = data.get("num_doc", "").strip()
    if not tipo_doc or not num_doc:
        return jsonify({"error": "Campos incompletos"}), 400
    PACIENTES_EPS = cargar_pacientes("data/bd_coosalud.txt")
    clave = f"{tipo_doc}|{num_doc}"
    paciente = PACIENTES_EPS.get(clave)
    if paciente:
        nombres = " ".join(filter(None, [paciente["Primer_Nombre"], paciente["Segundo_Nombre"]]))
        apellidos = " ".join(filter(None, [paciente["Primer_Apellido"], paciente["Segundo_Apellido"]]))
        return jsonify({
            "existe": True,
            "nombre_completo": f"{nombres} {apellidos}".strip(),
            "fecha_nacimiento": paciente["Fecha_Nacimiento"],
            "edad": paciente["Edad"],
            "genero": paciente["Genero"],
            "tipo_doc": tipo_doc,
            "num_doc": num_doc,
            "REGIMEN": paciente["REGIMEN"],
            "TIPO_AFILIADO": paciente["TIPO_AFILIADO"],
            "CODIGO_EPS": paciente["CODIGO_EPS"],
            "DEPARTAMENTO": paciente["DEPARTAMENTO"],
            "CIUDAD": paciente["CIUDAD"],
            "ZONA": paciente["ZONA"],
            "ESTADO": paciente["ESTADO"],
        })
    else:
        return jsonify({"existe": False, "error": "Paciente no encontrado o no activo."})
@app.route("/validar-dx", methods=["POST"])
@login_requerido
def validar_dx():
    dx = request.get_json().get("dx", "").strip().upper()
    DX_AUTORIZADOS = cargar_dx("data/dx.txt")
    desc = DX_AUTORIZADOS.get(dx)
    return jsonify({"valido": bool(desc), "descripcion": desc or "No autorizado", "dx_codigo": dx})
@app.route("/validar-servicio", methods=["POST"])
@login_requerido
def validar_servicio():
    cups = request.get_json().get("cups", "").strip()
    SERVICIOS_AUTORIZADOS = cargar_servicios("data/servicios.txt")
    nombre = SERVICIOS_AUTORIZADOS.get(cups)
    es_terapia = cups in TERAPIAS_REPETITIVAS
    return jsonify({
        "valido": bool(nombre),
        "nombre": nombre or "No autorizado",
        "cups": cups,
        "es_terapia": es_terapia
    })
@app.route("/progreso-terapia", methods=["POST"])
@login_requerido
def progreso_terapia():
    data = request.get_json()
    tipo_doc = data.get("tipo_doc")
    num_doc = data.get("num_doc")
    cups = data.get("cups")
    total_autorizado = int(data.get("total", 0))
    if not tipo_doc or not num_doc or not cups:
        return jsonify({"error": "Faltan datos"}), 400
    realizadas = contar_sesiones_realizadas(tipo_doc, num_doc, cups)
    pendientes = max(0, total_autorizado - realizadas)
    return jsonify({
        "realizadas": realizadas,
        "pendientes": pendientes,
        "total": total_autorizado
    })
# ✅ CORREGIDO: Validación de paciente duplicado (MODIFICADA - Deglución = Individual como Autismo)
@app.route("/guardar-cita", methods=["POST"])
@login_requerido
def guardar_cita():
    data = request.get_json()
    required = ["paciente", "dx", "celular", "email", "servicio", "profesional", "fecha", "hora"]
    if not all(k in data for k in required):
        return jsonify({"exito": False, "mensaje": "Faltan datos."})
    dx_paciente = data["dx"]["dx_codigo"].upper()
    cups_paciente = data["servicio"]["cups"]
    es_autismo_nuevo = dx_paciente == "F840"
    es_deglucion_nuevo = cups_paciente == "937203"  # ✅ DEGLUCIÓN = INDIVIDUAL
    es_individual = es_autismo_nuevo or es_deglucion_nuevo  # ✅ AMBOS REQUIEREN INDIVIDUAL
    profesional = data["profesional"]
    fecha = data["fecha"]
    hora = data["hora"]
    paciente_num_doc = data["paciente"]["num_doc"]
    paciente_tipo_doc = data["paciente"]["tipo_doc"]
    clave = f"{profesional}|{fecha}|{hora}"
    AGENDA_ACTUAL = cargar_agenda_desde_csv_filtrada()
    citas_existentes = AGENDA_ACTUAL.get(clave, [])
    tiene_autismo_existente = any(cita.get("_es_autismo") for cita in citas_existentes)
    tiene_deglucion_existente = any(cita.get("_es_deglucion") for cita in citas_existentes)  # ✅ AGREGADO
    tiene_individual_existente = tiene_autismo_existente or tiene_deglucion_existente  # ✅ VERIFICAR AMBOS
    
    # ✅ REGLA: Si hay autismo o deglución existente, no permitir otro paciente
    if tiene_individual_existente and not es_individual:
        return jsonify({"exito": False, "mensaje": "❌ Este horario ya tiene un paciente que requiere sesión individual (Autismo F840 o Deglución 937203). No se permiten sesiones compartidas."})
    
    # ✅ REGLA: Si el nuevo paciente es autismo o deglución, requiere horario vacío
    if es_individual and len(citas_existentes) > 0:
        return jsonify({"exito": False, "mensaje": "❌ Pacientes con autismo (F840) o deglución (937203) requieren sesión individual. Elige un horario vacío."})
    
    # ✅ REGLA: Si no es individual, máximo 2 pacientes
    if not es_individual and len(citas_existentes) >= 2:
        return jsonify({"exito": False, "mensaje": "❌ Horario lleno (máximo 2 pacientes)."})
    
    # ✅ NUEVA VALIDACIÓN: Paciente no puede estar en dos lugares a la misma hora
    if os.path.exists("data/agenda.csv"):
        with open("data/agenda.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if (row.get("Num_Doc") == paciente_num_doc and
                    row.get("Tipo_Doc") == paciente_tipo_doc and
                    row.get("Fecha") == fecha and
                    row.get("Hora") == hora and
                    row.get("Estado", "activa") == "activa" and
                    row.get("Profesional") != profesional):
                    return jsonify({
                        "exito": False,
                        "mensaje": f"❌ CONFLICTO: Este paciente ya tiene una cita agendada el {fecha} a las {hora} con el profesional '{row.get('Profesional')}'. No se puede agendar con dos profesionales a la misma hora."
                    })
    paciente_data = data["paciente"]
    cita = {
        "Tipo_Doc": paciente_data["tipo_doc"],
        "Num_Doc": paciente_data["num_doc"],
        "Nombre_Completo": paciente_data["nombre_completo"],
        "Edad": str(paciente_data.get("edad") or ""),
        "Genero": paciente_data["genero"],
        "REGIMEN": paciente_data.get("REGIMEN", ""),
        "TIPO_AFILIADO": paciente_data.get("TIPO_AFILIADO", ""),
        "CODIGO_EPS": paciente_data.get("CODIGO_EPS", ""),
        "DEPARTAMENTO": paciente_data.get("DEPARTAMENTO", ""),
        "CIUDAD": paciente_data.get("CIUDAD", ""),
        "ZONA": paciente_data.get("ZONA", ""),
        "ESTADO": paciente_data.get("ESTADO", ""),
        "Celular": data["celular"],
        "Email": data["email"],
        "Dx_Codigo": data["dx"]["dx_codigo"],
        "Dx_Descripcion": data["dx"]["descripcion"],
        "CUPS": data["servicio"]["cups"],
        "Servicio": data["servicio"]["nombre"],
        "Profesional": profesional,
        "Fecha": fecha,
        "Hora": hora,
        "Cantidad_Total": str(data["servicio"].get("cantidad_total", "")),
        "Frecuencia_Semanal": str(data["servicio"].get("frecuencia_semanal", "")),
        "Duracion_Meses": str(data["servicio"].get("duracion_meses", "")),
        "Observacion": data["servicio"].get("observacion", ""),
        "Fecha_Registro": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Estado": "activa"
    }
    cita_completa = {campo: cita.get(campo, "") for campo in CAMPOS_AGENDA}
    existe = os.path.exists("data/agenda.csv")
    with open("data/agenda.csv", "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CAMPOS_AGENDA)
        if not existe:
            writer.writeheader()
        writer.writerow(cita_completa)
    return jsonify({"exito": True, "mensaje": "Registro guardado correctamente."})
# ✅ CORREGIDO: Validación de paciente duplicado en bloques (MODIFICADA - Deglución = Individual como Autismo)
@app.route("/guardar-cita-bloque", methods=["POST"])
@login_requerido
def guardar_cita_bloque():
    data = request.get_json()
    required = ["paciente", "dx", "celular", "email", "servicio", "profesional", "fecha_inicio", "hora"]
    if not all(k in data for k in required):
        return jsonify({"exito": False, "mensaje": "Faltan datos."})
    paciente = data["paciente"]
    servicio = data["servicio"]
    profesional = data["profesional"]
    cantidad_total = int(servicio.get("cantidad_total", 1))
    frecuencia_semanal = int(servicio.get("frecuencia_semanal", 1))
    sesiones_por_visita = int(servicio.get("sesiones_por_visita", 1))
    duracion_meses = int(servicio.get("duracion_meses", 1))
    dx_paciente = data["dx"]["dx_codigo"].upper()
    cups_paciente = servicio.get("cups", "")
    es_autismo = dx_paciente == "F840"
    es_deglucion = cups_paciente == "937203"  # ✅ DEGLUCIÓN = INDIVIDUAL
    es_individual = es_autismo or es_deglucion  # ✅ AMBOS REQUIEREN INDIVIDUAL
    if cantidad_total % sesiones_por_visita != 0:
        return jsonify({"exito": False, "mensaje": "❌ La cantidad total debe ser múltiplo de las sesiones por visita."})
    total_visitas = cantidad_total // sesiones_por_visita
    fechas = generar_fechas_terapia(data["fecha_inicio"], frecuencia_semanal, total_visitas)
    fechas = list(dict.fromkeys(fechas))
    if not fechas or len(fechas) != total_visitas:
        return jsonify({"exito": False, "mensaje": f"❌ No se pudieron generar {total_visitas} fechas válidas."})
    fechas_no_habiles = [f for f in fechas if not es_dia_habil(f)]
    if fechas_no_habiles:
        return jsonify({"exito": False, "mensaje": f"❌ Fechas no hábiles detectadas: {', '.join(fechas_no_habiles[:3])}"})
    agenda_actual = cargar_agenda_desde_csv_filtrada()
    conflictos = []
    for fecha in fechas:
        if data["hora"] not in TODOS_BLOQUES:
            conflictos.append(f"{fecha}: Hora inválida.")
            continue
        idx_inicio = TODOS_BLOQUES.index(data["hora"])
        if idx_inicio + sesiones_por_visita > len(TODOS_BLOQUES):
            conflictos.append(f"{fecha}: No hay suficientes bloques consecutivos desde {data['hora']}.")
            continue
        horas_necesarias = TODOS_BLOQUES[idx_inicio : idx_inicio + sesiones_por_visita]
        for hora in horas_necesarias:
            clave = f"{profesional}|{fecha}|{hora}"
            citas_existentes = agenda_actual.get(clave, [])
            tiene_autismo_existente = any(cita.get("_es_autismo") for cita in citas_existentes)
            tiene_deglucion_existente = any(cita.get("_es_deglucion") for cita in citas_existentes)  # ✅ AGREGADO
            tiene_individual_existente = tiene_autismo_existente or tiene_deglucion_existente  # ✅ VERIFICAR AMBOS
            ocupados = len(citas_existentes)
            # ✅ VALIDACIÓN DE PACIENTE DUPLICADO
            if os.path.exists("data/agenda.csv"):
                with open("data/agenda.csv", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if (row.get("Num_Doc") == paciente["num_doc"] and
                            row.get("Tipo_Doc") == paciente["tipo_doc"] and
                            row.get("Fecha") == fecha and
                            row.get("Hora") == hora and
                            row.get("Estado", "activa") == "activa" and
                            row.get("Profesional") != profesional):
                            conflictos.append(f"{fecha} {hora} → Paciente ya agendado con '{row.get('Profesional')}'")
                            break
            if es_individual:  # ✅ AMBOS (autismo y deglución) requieren individual
                if ocupados > 0:
                    conflictos.append(f"{fecha} {hora} → Ya hay cita(s). Paciente requiere sesión individual (Autismo/Deglución).")
            else:
                if tiene_individual_existente:  # ✅ VERIFICAR AMBOS
                    conflictos.append(f"{fecha} {hora} → Ya hay paciente que requiere sesión individual (Autismo/Deglución). Sesión exclusiva.")
                elif ocupados >= 2:
                    conflictos.append(f"{fecha} {hora} → Horario lleno (máx. 2 pacientes).")
    if conflictos:
        mensaje = "❌ **Agendamiento fallido**: no se guardó ninguna cita.\nConflictos:\n" + "\n".join(conflictos[:10])
        if len(conflictos) > 10:
            mensaje += f"\n... y {len(conflictos) - 10} más."
        return jsonify({"exito": False, "mensaje": mensaje, "sugerir_busqueda": True})
    #eliminar_citas_bloque_existente(paciente["tipo_doc"], paciente["num_doc"], servicio["cups"], profesional)
    with open("data/agenda.csv", "a", newline="", encoding="utf-8") as f:
        writer = None
        for fecha in fechas:
            idx_inicio = TODOS_BLOQUES.index(data["hora"])
            horas_necesarias = TODOS_BLOQUES[idx_inicio : idx_inicio + sesiones_por_visita]
            for hora in horas_necesarias:
                cita = {
                    "Tipo_Doc": paciente["tipo_doc"],
                    "Num_Doc": paciente["num_doc"],
                    "Nombre_Completo": paciente["nombre_completo"],
                    "Edad": str(paciente.get("edad") or ""),
                    "Genero": paciente["genero"],
                    "REGIMEN": paciente.get("REGIMEN", ""),
                    "TIPO_AFILIADO": paciente.get("TIPO_AFILIADO", ""),
                    "CODIGO_EPS": paciente.get("CODIGO_EPS", ""),
                    "DEPARTAMENTO": paciente.get("DEPARTAMENTO", ""),
                    "CIUDAD": paciente.get("CIUDAD", ""),
                    "ZONA": paciente.get("ZONA", ""),
                    "ESTADO": paciente.get("ESTADO", ""),
                    "Celular": data["celular"],
                    "Email": data["email"],
                    "Dx_Codigo": data["dx"]["dx_codigo"],
                    "Dx_Descripcion": data["dx"]["descripcion"],
                    "CUPS": servicio["cups"],
                    "Servicio": servicio["nombre"],
                    "Profesional": profesional,
                    "Fecha": fecha,
                    "Hora": hora,
                    "Cantidad_Total": str(cantidad_total),
                    "Frecuencia_Semanal": str(frecuencia_semanal),
                    "Duracion_Meses": str(duracion_meses),
                    "Observacion": servicio.get("observacion", ""),
                    "Fecha_Registro": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "Estado": "activa"
                }
                cita_completa = {campo: cita.get(campo, "") for campo in CAMPOS_AGENDA}
                if writer is None:
                    existe = os.path.exists("data/agenda.csv") and os.path.getsize("data/agenda.csv") > 0
                    writer = csv.DictWriter(f, fieldnames=CAMPOS_AGENDA)
                    if not existe:
                        writer.writeheader()
                writer.writerow(cita_completa)
    msg = f"✅ Bloque agendado: {cantidad_total} sesiones en {total_visitas} visitas ({sesiones_por_visita} por visita)."
    return jsonify({"exito": True, "mensaje": msg})
@app.route("/cancelar-cita", methods=["POST"])
@login_requerido
def cancelar_cita():
    data = request.get_json()
    required = ["tipo_doc", "num_doc", "profesional", "fecha", "hora", "cups"]
    if not all(k in data for k in required):
        return jsonify({"exito": False, "mensaje": "Faltan datos."})
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"exito": False, "mensaje": "No existe agenda registrada."})
    registros_actualizados = []
    cita_encontrada = False
    fieldnames = None
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or CAMPOS_AGENDA
        for row in reader:
            if (row.get("Tipo_Doc") == data["tipo_doc"] and
                row.get("Num_Doc") == data["num_doc"] and
                row.get("Profesional") == data["profesional"] and
                row.get("Fecha") == data["fecha"] and
                row.get("Hora") == data["hora"] and
                row.get("CUPS") == data["cups"] and
                row.get("Estado", "activa") == "activa"):
                row["Estado"] = "cancelada"
                cita_encontrada = True
            elif "Estado" not in row:
                row["Estado"] = "activa"
            registros_actualizados.append(row)
    if not cita_encontrada:
        return jsonify({"exito": False, "mensaje": "Cita no encontrada o ya cancelada."})
    with open("data/agenda.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(registros_actualizados)
    return jsonify({"exito": True, "mensaje": "Cita cancelada correctamente."})
@app.route("/marcar-no-asistio", methods=["POST"])
@login_requerido
def marcar_no_asistio():
    data = request.get_json()
    required = ["tipo_doc", "num_doc", "profesional", "fecha", "hora", "cups"]
    if not all(k in data for k in required):
        return jsonify({"exito": False, "mensaje": "Faltan datos."})
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"exito": False, "mensaje": "No existe agenda registrada."})
    registros_actualizados = []
    cita_encontrada = False
    fieldnames = None
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or CAMPOS_AGENDA
        for row in reader:
            if (row.get("Tipo_Doc") == data["tipo_doc"] and
                row.get("Num_Doc") == data["num_doc"] and
                row.get("Profesional") == data["profesional"] and
                row.get("Fecha") == data["fecha"] and
                row.get("Hora") == data["hora"] and
                row.get("CUPS") == data["cups"] and
                row.get("Estado", "activa") == "activa"):
                row["Estado"] = "no_asistio"
                cita_encontrada = True
            elif "Estado" not in row:
                row["Estado"] = "activa"
            registros_actualizados.append(row)
    if not cita_encontrada:
        return jsonify({"exito": False, "mensaje": "Cita no encontrada o ya finalizada."})
    with open("data/agenda.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(registros_actualizados)
    return jsonify({"exito": True, "mensaje": "Cita marcada como 'no asistió'."})
@app.route("/detalles-horario", methods=["POST"])
@login_requerido
def detalles_horario():
    data = request.get_json()
    profesional = data.get("profesional")
    fecha = data.get("fecha")
    hora = data.get("hora")
    if not profesional or not fecha or not hora:
        return jsonify({"error": "Faltan datos"}), 400
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"citas": []})
    citas_en_horario = []
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if (row.get("Profesional") == profesional and
                row.get("Fecha") == fecha and
                row.get("Hora") == hora and
                row.get("Estado", "activa") == "activa"):
                citas_en_horario.append(row)
    return jsonify({"citas": citas_en_horario})
@app.route("/detalles-profesional", methods=["POST"])
@login_requerido
def detalles_profesional():
    data = request.get_json()
    profesional = data.get("profesional")
    fecha = data.get("fecha")
    if not profesional or not fecha:
        return jsonify({"error": "Faltan datos"}), 400
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"citas": []})
    citas_profesional = []
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if (row.get("Profesional") == profesional and
                row.get("Fecha") == fecha and
                row.get("Estado", "activa") == "activa"):
                citas_profesional.append(row)
    return jsonify({"citas": citas_profesional})
@app.route("/recordatorio-whatsapp", methods=["POST"])
@login_requerido
def recordatorio_whatsapp():
    data = request.get_json()
    required = ["tipo_doc", "num_doc", "profesional", "fecha", "hora", "cups"]
    if not all(k in data for k in required):
        return jsonify({"exito": False, "mensaje": "Faltan datos."})
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"exito": False, "mensaje": "No hay agenda registrada."})
    paciente_info = None
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if (row.get("Tipo_Doc") == data["tipo_doc"] and
                row.get("Num_Doc") == data["num_doc"] and
                row.get("Profesional") == data["profesional"] and
                row.get("Fecha") == data["fecha"] and
                row.get("Hora") == data["hora"] and
                row.get("CUPS") == data["cups"] and
                row.get("Estado", "activa") == "activa"):
                paciente_info = row
                break
    if not paciente_info:
        return jsonify({"exito": False, "mensaje": "Cita no encontrada o cancelada."})
    celular = paciente_info.get("Celular", "").strip()
    if not celular:
        return jsonify({"exito": False, "mensaje": "El paciente no tiene número de celular registrado."})
    celular_limpio = "".join(filter(str.isdigit, celular))
    if celular_limpio and not celular_limpio.startswith("57"):
        celular_limpio = "57" + celular_limpio
    fecha_humana = datetime.strptime(data["fecha"], "%Y-%m-%d").strftime("%d/%m/%Y")
    mensaje = (
        f"🔔 *Recordatorio de Cita Médica*\n"
        f"Estimado(a) *{paciente_info['Nombre_Completo']}*,\n"
        f"Le recordamos su cita programada:\n"
        f"📅 Fecha: {fecha_humana}\n"
        f"🕗 Hora: {data['hora']}\n"
        f"👤 Profesional: {data['profesional']}\n"
        f"⚕️ Servicio: {paciente_info['Servicio']} ({data['cups']})\n"
        f"📍 NEURODESARROLLO\n"
        f"Cra 36 #54-50 Cabecera Bucaramanga"
    )
    return jsonify({"exito": True, "mensaje": mensaje, "celular": celular_limpio})
@app.route("/recordatorios-manana", methods=["GET"])
@login_requerido
def recordatorios_manana():
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"exito": False, "mensaje": "No hay agenda registrada."})
    manana = (date.today() + timedelta(days=1)).isoformat()
    recordatorios = []
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("Fecha") == manana and row.get("Estado", "activa") == "activa":
                celular = row.get("Celular", "").strip()
                if not celular:
                    continue
                celular_limpio = "".join(filter(str.isdigit, celular))
                if celular_limpio and not celular_limpio.startswith("57"):
                    celular_limpio = "57" + celular_limpio
                fecha_humana = datetime.strptime(manana, "%Y-%m-%d").strftime("%d/%m/%Y")
                mensaje = (
                    f"🔔 *Recordatorio de Cita Médica*\n"
                    f"Estimado(a) *{row['Nombre_Completo']}*,\n"
                    f"Le recordamos su cita programada para **mañana**:\n"
                    f"📅 Fecha: {fecha_humana}\n"
                    f"🕗 Hora: {row['Hora']}\n"
                    f"👤 Profesional: {row['Profesional']}\n"
                    f"⚕️ Servicio: {row['Servicio']} ({row['CUPS']})\n"
                    f"📍 NEURODESARROLLO\n"
                    f"Cra 36 #54-50 Cabecera Bucaramanga"
                )
                recordatorios.append({"nombre": row["Nombre_Completo"], "celular": celular_limpio, "mensaje": mensaje})
    if not recordatorios:
        return jsonify({"exito": False, "mensaje": "No hay citas programadas para mañana con celular registrado."})
    return jsonify({"exito": True, "fecha": manana, "recordatorios": recordatorios})
def leer_txt_como_lista(filepath):
    if not os.path.exists(filepath):
        return []
    with open(filepath, encoding="utf-8") as f:
        lines = f.readlines()
        return [line.strip() for line in lines if line.strip()]
def escribir_txt_como_lista(filepath, lines):
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
@app.route("/admin")
@login_requerido
def admin_panel():
    if session.get('rol') != 'admin':
        return redirect(url_for('index'))
    return render_template("admin.html")
@app.route("/admin/dx/lista")
@login_requerido
def lista_dx():
    if session.get('rol') != 'admin':
        return jsonify([])
    return jsonify(leer_txt_como_lista("data/dx.txt"))
@app.route("/admin/servicios/lista")
@login_requerido
def lista_servicios():
    if session.get('rol') != 'admin':
        return jsonify([])
    return jsonify(leer_txt_como_lista("data/servicios.txt"))
@app.route("/admin/profesionales/lista")
@login_requerido
def lista_profesionales():
    if session.get('rol') != 'admin':
        return jsonify([])
    return jsonify(leer_txt_como_lista("data/Profesionales.txt"))
@app.route("/admin/dx/agregar", methods=["POST"])
@login_requerido
def agregar_dx():
    if session.get('rol') != 'admin':
        return jsonify({"exito": False, "mensaje": "Acceso denegado."})
    data = request.get_json()
    codigo = data.get("codigo", "").strip()
    nombre = data.get("nombre", "").strip()
    if not codigo or not nombre:
        return jsonify({"exito": False, "mensaje": "Faltan código o nombre."})
    lines = leer_txt_como_lista("data/dx.txt")
    if not any(codigo == line.split("|")[0] for line in lines[1:]):
        lines.append(f"{codigo}|{nombre}")
        escribir_txt_como_lista("data/dx.txt", lines)
        return jsonify({"exito": True, "mensaje": "Dx agregado."})
    return jsonify({"exito": False, "mensaje": "Dx ya existe."})
@app.route("/admin/servicios/agregar", methods=["POST"])
@login_requerido
def agregar_servicio():
    if session.get('rol') != 'admin':
        return jsonify({"exito": False, "mensaje": "Acceso denegado."})
    data = request.get_json()
    cups = data.get("cups", "").strip()
    hom = data.get("hom", "N/A").strip()
    nombre = data.get("nombre", "").strip()
    if not cups or not nombre:
        return jsonify({"exito": False, "mensaje": "Faltan CUPS o nombre."})
    lines = leer_txt_como_lista("data/servicios.txt")
    if not any(cups == line.split("|")[0] for line in lines[1:]):
        lines.append(f"{cups}|{hom}|{nombre}")
        escribir_txt_como_lista("data/servicios.txt", lines)
        return jsonify({"exito": True, "mensaje": "Servicio agregado."})
    return jsonify({"exito": False, "mensaje": "Servicio ya existe."})
@app.route("/admin/profesionales/agregar", methods=["POST"])
@login_requerido
def agregar_profesional():
    if session.get('rol') != 'admin':
        return jsonify({"exito": False, "mensaje": "Acceso denegado."})
    data = request.get_json()
    nombre = data.get("nombre", "").strip()
    esp1 = data.get("esp1", "").strip()
    esp2 = data.get("esp2", "").strip()
    cel = data.get("cel", "3").strip()
    if not nombre:
        return jsonify({"exito": False, "mensaje": "Falta nombre del profesional."})
    lines = leer_txt_como_lista("data/Profesionales.txt")
    if not any(nombre == line.split("|")[0] for line in lines[1:]):
        lines.append(f"{nombre}|{esp1}|{esp2}|{cel}")
        escribir_txt_como_lista("data/Profesionales.txt", lines)
        return jsonify({"exito": True, "mensaje": "Profesional agregado."})
    return jsonify({"exito": False, "mensaje": "Profesional ya existe."})
@app.route("/admin/dx/eliminar", methods=["POST"])
@login_requerido
def eliminar_dx():
    if session.get('rol') != 'admin':
        return jsonify({"exito": False, "mensaje": "Acceso denegado."})
    data = request.get_json()
    codigo = data.get("codigo", "").strip()
    if not codigo:
        return jsonify({"exito": False, "mensaje": "Falta código."})
    lines = leer_txt_como_lista("data/dx.txt")
    nuevas = [lines[0]]
    for line in lines[1:]:
        if line.split("|")[0] != codigo:
            nuevas.append(line)
    if len(nuevas) > 1:
        escribir_txt_como_lista("data/dx.txt", nuevas)
        return jsonify({"exito": True, "mensaje": "Dx eliminado."})
    return jsonify({"exito": False, "mensaje": "Dx no encontrado."})
@app.route("/admin/servicios/eliminar", methods=["POST"])
@login_requerido
def eliminar_servicio():
    if session.get('rol') != 'admin':
        return jsonify({"exito": False, "mensaje": "Acceso denegado."})
    data = request.get_json()
    cups = data.get("cups", "").strip()
    if not cups:
        return jsonify({"exito": False, "mensaje": "Falta CUPS."})
    lines = leer_txt_como_lista("data/servicios.txt")
    nuevas = [lines[0]]
    for line in lines[1:]:
        if line.split("|")[0] != cups:
            nuevas.append(line)
    if len(nuevas) > 1:
        escribir_txt_como_lista("data/servicios.txt", nuevas)
        return jsonify({"exito": True, "mensaje": "Servicio eliminado."})
    return jsonify({"exito": False, "mensaje": "Servicio no encontrado."})
@app.route("/admin/profesionales/eliminar", methods=["POST"])
@login_requerido
def eliminar_profesional():
    if session.get('rol') != 'admin':
        return jsonify({"exito": False, "mensaje": "Acceso denegado."})
    data = request.get_json()
    nombre = data.get("nombre", "").strip()
    if not nombre:
        return jsonify({"exito": False, "mensaje": "Falta nombre."})
    lines = leer_txt_como_lista("data/Profesionales.txt")
    nuevas = [lines[0]]
    for line in lines[1:]:
        if line.split("|")[0] != nombre:
            nuevas.append(line)
    if len(nuevas) > 1:
        escribir_txt_como_lista("data/Profesionales.txt", nuevas)
        return jsonify({"exito": True, "mensaje": "Profesional eliminado."})
    return jsonify({"exito": False, "mensaje": "Profesional no encontrado."})
@app.route("/admin/profesionales/editar/celular", methods=["POST"])
@login_requerido
def editar_celular_profesional():
    if session.get('rol') != 'admin':
        return jsonify({"exito": False, "mensaje": "Acceso denegado."})
    data = request.get_json()
    nombre = data.get("nombre", "").strip()
    celular = data.get("celular", "3").strip()
    if not nombre:
        return jsonify({"exito": False, "mensaje": "Falta nombre."})
    lines = leer_txt_como_lista("data/Profesionales.txt")
    nuevas = [lines[0]]
    encontrado = False
    for line in lines[1:]:
        partes = line.split("|")
        if partes[0].strip() == nombre.strip():
            while len(partes) < 4:
                partes.append("3")
            partes[3] = celular
            nuevas.append("|".join(partes))
            encontrado = True
        else:
            nuevas.append(line)
    if encontrado:
        escribir_txt_como_lista("data/Profesionales.txt", nuevas)
        return jsonify({"exito": True, "mensaje": "Celular actualizado."})
    return jsonify({"exito": False, "mensaje": f"Profesional '{nombre}' no encontrado."})
@app.route("/whatsapp/mensaje", methods=["POST"])
@login_requerido
def generar_mensaje_whatsapp():
    data = request.get_json()
    profesional = data.get("profesional", "").strip()
    fecha = data.get("fecha", "").strip()
    if not profesional or not fecha:
        return jsonify({"exito": False, "mensaje": "Faltan profesional o fecha."})
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"exito": False, "mensaje": "No hay agenda registrada."})
    pacientes = []
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if (row.get("Profesional") == profesional and
                row.get("Fecha") == fecha and
                row.get("Estado", "activa") == "activa"):
                pacientes.append({"hora": row["Hora"], "nombre": row["Nombre_Completo"], "cups": row["CUPS"]})
    if not pacientes:
        return jsonify({"exito": False, "mensaje": "No hay citas para este profesional en la fecha indicada."})
    mensaje = f"📅 *Agenda - {fecha}*\n*Profesional:* {profesional}\n"
    bloques = defaultdict(list)
    for p in pacientes:
        bloques[p["hora"]].append(p["nombre"])
    for hora in sorted(bloques.keys()):
        nombres = ", ".join(bloques[hora])
        mensaje += f"🕗 {hora}: {nombres}\n"
    mensaje += f"\nTotal de citas: {len(pacientes)}"
    celular = obtener_celular_profesional(profesional)
    celular = "".join(filter(str.isdigit, celular))
    if celular and not celular.startswith("57"):
        celular = "57" + celular
    return jsonify({"exito": True, "mensaje": mensaje, "celular": celular})
@app.route("/whatsapp/agenda-diaria", methods=["POST"])
@login_requerido
def enviar_agendas_diarias():
    data = request.get_json()
    fecha = data.get("fecha", "").strip()
    if not fecha:
        return jsonify({"exito": False, "mensaje": "Falta la fecha."})
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"exito": False, "mensaje": "No hay agenda registrada."})
    agenda_dia = defaultdict(list)
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("Fecha") == fecha and row.get("Estado", "activa") == "activa":
                agenda_dia[row["Profesional"]].append({"hora": row["Hora"], "nombre": row["Nombre_Completo"]})
    if not agenda_dia:
        return jsonify({"exito": False, "mensaje": "No hay citas programadas para esta fecha."})
    celulares = cargar_celulares_profesionales()
    profesionales_mensajes = []
    for profesional, citas in agenda_dia.items():
        mensaje = f"📅 *Agenda - {fecha}*\n*Profesional:* {profesional}\n"
        bloques = defaultdict(list)
        for cita in citas:
            bloques[cita["hora"]].append(cita["nombre"])
        for hora in sorted(bloques.keys()):
            nombres = ", ".join(bloques[hora])
            mensaje += f"🕗 {hora}: {nombres}\n"
        mensaje += f"\nTotal de citas: {len(citas)}"
        celular = celulares.get(profesional, "")
        if celular:
            profesionales_mensajes.append({"profesional": profesional, "mensaje": mensaje, "celular": celular})
    if not profesionales_mensajes:
        return jsonify({"exito": False, "mensaje": "Ningún profesional tiene número de celular válido."})
    return jsonify({"exito": True, "fecha": fecha, "mensajes": profesionales_mensajes})
@app.route("/reportes")
@login_requerido
def reportes_panel():
    return render_template("reportes.html")

@app.route("/generar-reporte", methods=["POST"])
@login_requerido
def generar_reporte():
    data = request.get_json()
    tipo = data.get("tipo")
    fecha_inicio = data.get("fecha_inicio")
    fecha_fin = data.get("fecha_fin")
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"error": "No hay datos para generar reportes."})
    try:
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        if fi > ff:
            return jsonify({"error": "La fecha de inicio no puede ser mayor que la final."})
    except:
        return jsonify({"error": "Formato de fecha inválido."})
    registros = []
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                fecha_cita = datetime.strptime(row["Fecha"], "%Y-%m-%d").date()
                if fi <= fecha_cita <= ff:
                    registros.append(row)
            except:
                continue
    if not registros:
        return jsonify({"error": "No hay registros en el rango de fechas seleccionado."})
    rol = session.get('rol', 'admin')
    profesional_logueado = session.get('usuario', '')
    if rol == 'profesional':
        registros = [r for r in registros if r.get("Profesional") == profesional_logueado]
    if tipo == "reporte_diagnosticos":
        registros_activos = [r for r in registros if r.get("Estado", "activa") == "activa"]
        diag_data = defaultdict(lambda: {"citas": 0, "pacientes": set()})
        for r in registros_activos:
            dx_codigo = r.get("Dx_Codigo", "SIN_DX")
            dx_desc = r.get("Dx_Descripcion", "No especificado")
            clave = f"{dx_codigo} - {dx_desc}"
            diag_data[clave]["citas"] += 1
            diag_data[clave]["pacientes"].add(f"{r['Tipo_Doc']}-{r['Num_Doc']}")
        html = "<h3>📊 Reporte de Diagnósticos</h3>"
        html += f"<p><strong>Período:</strong> {fecha_inicio} a {fecha_fin}</p>"
        html += "<table border='1' style='width:100%; border-collapse:collapse; font-size:13px;'>"
        html += "<tr><th>Código Dx</th><th>Diagnóstico</th><th>Total Citas</th><th>Pacientes Únicos</th></tr>"
        for dx_completo, datos in sorted(diag_data.items(), key=lambda x: x[1]['citas'], reverse=True):
            partes = dx_completo.split(" - ", 1)
            codigo_dx = partes[0]
            desc_dx = partes[1] if len(partes) > 1 else "Sin descripción"
            html += f"<tr><td>{codigo_dx}</td><td>{desc_dx}</td><td>{datos['citas']}</td><td>{len(datos['pacientes'])}</td></tr>"
        html += "</table>"
        return jsonify({"html": html})
    elif tipo == "reporte_general":
        total_citas = len([r for r in registros if r.get("Estado", "activa") == "activa"])
        pacientes_unicos = len(set(r["Num_Doc"] for r in registros if r.get("Estado", "activa") == "activa"))
        profesionales_unicos = len(set(r["Profesional"] for r in registros if r.get("Estado", "activa") == "activa"))
        servicios = defaultdict(int)
        for r in registros:
            servicios[r["CUPS"]] += 1
        nombres_serv = {row["CUPS_ISS"]: row["SERVICIO"] for row in csv.DictReader(open("data/servicios.txt", encoding="utf-8"), delimiter="|")} if os.path.exists("data/servicios.txt") else {}
        html = "<h3>📊 Reporte General</h3>"
        html += f"<p><strong>Período:</strong> {fecha_inicio} a {fecha_fin}</p>"
        html += f"<p><strong>Total de Citas Activas:</strong> {total_citas}</p>"
        html += f"<p><strong>Pacientes Atendidos:</strong> {pacientes_unicos}</p>"
        html += f"<p><strong>Profesionales Activos:</strong> {profesionales_unicos}</p>"
        html += "<h4>Servicios (CUPS) más utilizados:</h4>"
        html += "<table border='1' style='width:100%; border-collapse:collapse;'>"
        html += "<tr><th>CUPS</th><th>Nombre del Servicio</th><th>Cantidad</th></tr>"
        for cups, count in sorted(servicios.items(), key=lambda x: x[1], reverse=True):
            nombre = nombres_serv.get(cups, cups)
            html += f"<tr><td>{cups}</td><td>{nombre}</td><td>{count}</td></tr>"
        html += "</table>"
        return jsonify({"html": html})
    elif tipo == "reporte_por_paciente":
        tipo_doc_filtro = data.get("tipo_doc_paciente", "").strip().upper()
        num_doc_filtro = data.get("num_doc_paciente", "").strip()
        if not tipo_doc_filtro or not num_doc_filtro:
            return jsonify({"error": "Debe ingresar tipo y número de documento."})
        citas_paciente = [r for r in registros if r.get("Tipo_Doc") == tipo_doc_filtro and r.get("Num_Doc") == num_doc_filtro]
        if not citas_paciente:
            return jsonify({"error": f"No se encontraron citas para {tipo_doc_filtro}-{num_doc_filtro} en el rango seleccionado."})
        citas_paciente.sort(key=lambda x: x.get("Fecha", ""), reverse=True)
        html = f"<h3>🔍 Citas para {tipo_doc_filtro}-{num_doc_filtro}</h3>"
        if citas_paciente:
            html += f"<p><strong>Paciente:</strong> {citas_paciente[0].get('Nombre_Completo', 'Desconocido')}</p>"
            html += "<table border='1' style='width:100%; border-collapse:collapse; font-size:13px;'>"
            html += "<tr><th>Fecha</th><th>Hora</th><th>Profesional</th><th>Servicio (CUPS)</th><th>Estado</th><th>Dx</th><th>Teléfono</th></tr>"
            for r in citas_paciente:
                estado = r.get("Estado", "activa")
                color = "#28a745" if estado == "activa" else "#dc3545" if estado == "cancelada" else "#6f42c1" if estado == "no_asistio" else "#6c757d"
                estado_label = "Programada" if estado == "activa" else "Cancelada" if estado == "cancelada" else "No Asistió" if estado == "no_asistio" else estado
                html += f"<tr><td>{r.get('Fecha', '')}</td><td>{r.get('Hora', '')}</td><td>{r.get('Profesional', '')}</td><td>{r.get('Servicio', '')} ({r.get('CUPS', '')})</td><td style='color:{color}; font-weight:bold;'>{estado_label}</td><td>{r.get('Dx_Codigo', '')}</td><td>{r.get('Celular', '')}</td></tr>"
            html += "</table>"
        return jsonify({"html": html})
    elif tipo == "reporte_canceladas":
        canceladas = [r for r in registros if r.get("Estado") == "cancelada"]
        if not canceladas:
            return jsonify({"error": "No hay citas canceladas en el rango seleccionado."})
        html = "<h3>❌ Citas Canceladas</h3>"
        html += "<table border='1' style='width:100%; border-collapse:collapse; font-size:13px;'>"
        html += "<tr><th>Fecha</th><th>Hora</th><th>Paciente</th><th>Tipo/Doc</th><th>Servicio (CUPS)</th><th>Profesional</th><th>Dx</th><th>Teléfono</th><th>Fecha Registro</th></tr>"
        for r in canceladas:
            html += f"<tr><td>{r.get('Fecha', '')}</td><td>{r.get('Hora', '')}</td><td>{r.get('Nombre_Completo', '')}</td><td>{r.get('Tipo_Doc', '')}-{r.get('Num_Doc', '')}</td><td>{r.get('Servicio', '')} ({r.get('CUPS', '')})</td><td>{r.get('Profesional', '')}</td><td>{r.get('Dx_Codigo', '')}</td><td>{r.get('Celular', '')}</td><td>{r.get('Fecha_Registro', '')}</td></tr>"
        html += "</table>"
        return jsonify({"html": html})
    elif tipo == "reporte_no_asistieron":
        no_asistieron = [r for r in registros if r.get("Estado") == "no_asistio"]
        if not no_asistieron:
            return jsonify({"error": "No hay registros de inasistencia en el rango seleccionado."})
        html = "<h3>🟥 Citas con Inasistencia</h3>"
        html += "<table border='1' style='width:100%; border-collapse:collapse; font-size:13px;'>"
        html += "<tr><th>Fecha</th><th>Hora</th><th>Paciente</th><th>Tipo/Doc</th><th>Servicio (CUPS)</th><th>Profesional</th><th>Dx</th><th>Teléfono</th><th>Fecha Registro</th></tr>"
        for r in no_asistieron:
            html += f"<tr><td>{r.get('Fecha', '')}</td><td>{r.get('Hora', '')}</td><td>{r.get('Nombre_Completo', '')}</td><td>{r.get('Tipo_Doc', '')}-{r.get('Num_Doc', '')}</td><td>{r.get('Servicio', '')} ({r.get('CUPS', '')})</td><td>{r.get('Profesional', '')}</td><td>{r.get('Dx_Codigo', '')}</td><td>{r.get('Celular', '')}</td><td>{r.get('Fecha_Registro', '')}</td></tr>"
        html += "</table>"
        return jsonify({"html": html})
    
    elif tipo == "actividad_profesional":
        resumen = defaultdict(lambda: defaultdict(int))
        profesionales_pacientes = defaultdict(set)
        for r in registros:
            if r.get("Estado", "activa") == "activa":
                prof = r["Profesional"]
                servicio = f"{r['Servicio']} ({r['CUPS']})"
                resumen[prof][servicio] += 1
                profesionales_pacientes[prof].add(r["Num_Doc"])
        
        html = "<h3>Resumen de Actividades por Profesional y Servicio</h3>"
        html += f"<p><strong>Período:</strong> {fecha_inicio} a {fecha_fin}</p>"
        html += "<table border='1' style='width:100%; border-collapse:collapse; font-size:13px;'>"
        html += "<tr style='background:#004a80; color:white;'>"
        html += "<th style='padding:10px;'>Profesional</th>"
        html += "<th style='padding:10px;'>Servicio (CUPS)</th>"
        html += "<th style='padding:10px;'>Citas</th>"
        html += "<th style='padding:10px;'>Pacientes Atendidos</th>"
        html += "</tr>"
        
        for prof in sorted(resumen.keys()):
            servicios = resumen[prof]
            total_pacientes = len(profesionales_pacientes[prof])
            total_citas = sum(servicios.values())
            first = True
            for servicio, cant in sorted(servicios.items()):
                html += "<tr>"
                if first:
                    # ✅ PRIMERA FILA: Incluye Profesional y Pacientes con rowspan
                    html += f"<td rowspan='{len(servicios)}' style='padding:8px; font-weight:bold; background:#e7f3ff;'>{prof}</td>"
                    html += f"<td style='padding:8px;'>{servicio}</td>"
                    html += f"<td style='padding:8px; text-align:center;'>{cant}</td>"
                    html += f"<td rowspan='{len(servicios)}' style='padding:8px; text-align:center; background:#e7f3ff;'>{total_pacientes}</td>"
                    first = False
                else:
                    # ✅ FILAS SECUNDARIAS: Solo Servicio y Citas (las otras 2 columnas están cubiertas por rowspan)
                    html += f"<td style='padding:8px;'>{servicio}</td>"
                    html += f"<td style='padding:8px; text-align:center;'>{cant}</td>"
                html += "</tr>"
        
        html += "</table>"
        return jsonify({"html": html})
    
    elif tipo == "control_terapias":
        # === 🔍 DEBUG: Ver estados en registros ===
        estados_debug = defaultdict(int)
        for r in registros:
            estado_raw = r.get("Estado", "SIN_ESTADO")
            estados_debug[estado_raw] += 1
        print("\n" + "="*60)
        print("🔍 DEBUG control_terapias - ESTADOS EN REGISTROS")
        print("="*60)
        print(f"📊 Estados RAW: {dict(estados_debug)}")
        print("="*60 + "\n")
        # ================================================
        
        # ✅ CLAVE: Documento + Profesional + CUPS (para agrupar órdenes)
        control = defaultdict(lambda: {
            "autorizadas": 0, 
            "realizadas": 0,
            "no_asistio": 0,
            "nombre": "", 
            "servicio": "", 
            "documento": "",
            "profesional": "",
            "fecha_registro": ""
        })
        
        # === PASO 1: Contar AUTORIZADAS (TODAS las filas NO canceladas en TODO el CSV) ===
        with open("data/agenda.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                estado = row.get("Estado", "activa").strip().lower()
                
                # ✅ EXCLUIR CANCELADAS de autorizadas
                if estado == "cancelada":
                    continue
                
                profesional = row.get("Profesional", "Sin asignar")
                cups = row.get("CUPS", "SIN_CUPS")
                
                # ✅ CLAVE ÚNICA POR ORDEN
                clave = (row.get("Tipo_Doc"), row.get("Num_Doc"), profesional, cups)
                
                control[clave]["nombre"] = row.get("Nombre_Completo", "Desconocido")
                control[clave]["servicio"] = f"{cups} - {row.get('Servicio', '')}"
                control[clave]["documento"] = f"{row.get('Tipo_Doc', '')}-{row.get('Num_Doc', '')}"
                control[clave]["profesional"] = profesional
                control[clave]["fecha_registro"] = row.get("Fecha_Registro", "")[:10]
                
                # ✅ AUTORIZADAS = Contar TODAS las filas NO canceladas
                control[clave]["autorizadas"] += 1
        
        # === PASO 2: Contar REALIZADAS y NO ASISTIO (solo dentro del rango de fechas) ===
        for r in registros:
            estado = r.get("Estado", "activa").strip().lower()
            
            # ✅ EXCLUIR CANCELADAS
            if estado == "cancelada":
                continue
            
            profesional = r.get("Profesional", "Sin asignar")
            cups = r.get("CUPS", "SIN_CUPS")
            
            clave = (r.get("Tipo_Doc"), r.get("Num_Doc"), profesional, cups)
            
            # ✅ REALIZADAS = Solo 'activa' dentro del rango
            if estado == "activa":
                control[clave]["realizadas"] += 1
            
            # ✅ NO ASISTIO = Solo 'no_asistio' dentro del rango
            if estado == "no_asistio":
                control[clave]["no_asistio"] += 1
        
        # === 🔍 DEBUG: Ver resultados ===
        total_realizadas_debug = sum(d["realizadas"] for d in control.values())
        total_autorizadas_debug = sum(d["autorizadas"] for d in control.values())
        total_no_asistio_debug = sum(d["no_asistio"] for d in control.values())
        print("="*60)
        print("🔍 DEBUG control_terapias - RESULTADOS")
        print("="*60)
        print(f"📊 Total Autorizadas (no canceladas): {total_autorizadas_debug}")
        print(f"📊 Total Realizadas (activas en rango): {total_realizadas_debug}")
        print(f"📊 Total No Asistió (en rango): {total_no_asistio_debug}")
        print("="*60 + "\n")
        # ================================================
        
        # ✅ GENERAR HTML
        html = "<h3>📊 Control de Terapias (Autorizadas vs Realizadas)</h3>"
        html += f"<p><strong>Período:</strong> {fecha_inicio} a {fecha_fin}</p>"
        html += "<table border='1' style='width:100%; border-collapse:collapse; font-size:13px;'>"
        html += "<tr style='background:#004a80; color:white;'>"
        html += "<th>Paciente</th><th>Documento</th><th>Profesional</th><th>Servicio (CUPS)</th><th>Fecha Orden</th><th>Autorizadas</th><th>Realizadas</th><th>No Asistió</th><th>Pendientes</th><th>% Cumplimiento</th></tr>"
        
        for clave, datos in sorted(control.items(), key=lambda x: (x[1]["nombre"], x[1]["profesional"])):
            aut = datos["autorizadas"]
            real = datos["realizadas"]
            nas = datos["no_asistio"]
            pend = max(0, aut - real - nas)
            pct = round((real / aut * 100), 1) if aut > 0 else 0
            
            # ✅ FILTRO: NO mostrar órdenes donde todas están canceladas (autorizadas = 0)
            if aut == 0:
                continue
            
            color = "green" if pct >= 100 else "orange" if pct >= 70 else "red"
            
            html += f"<tr>"
            html += f"<td>{datos['nombre']}</td>"
            html += f"<td>{datos['documento']}</td>"
            html += f"<td>{datos['profesional']}</td>"
            html += f"<td>{datos['servicio']}</td>"
            html += f"<td>{datos['fecha_registro']}</td>"
            html += f"<td><strong>{aut}</strong></td>"
            html += f"<td><strong>{real}</strong></td>"
            html += f"<td style='color:#6f42c1; font-weight:bold;'>{nas}</td>"
            html += f"<td>{pend}</td>"
            html += f"<td style='color:{color}; font-weight:bold;'>{pct}%</td>"
            html += f"</tr>"
        
        html += "</table>"
        return jsonify({"html": html})
    # === REPORTE: Terapias por Paciente ===
    elif tipo == "terapias_paciente":
        resumen = defaultdict(lambda: {"citas": 0, "servicios": defaultdict(int), "profesionales": set()})
        for r in registros:
            if r.get("Estado", "activa") == "activa":
                paciente = f"{r['Nombre_Completo']} ({r['Tipo_Doc']}-{r['Num_Doc']})"
                resumen[paciente]["citas"] += 1
                resumen[paciente]["servicios"][r["Servicio"]] += 1
                resumen[paciente]["profesionales"].add(r["Profesional"])
        html = "<h3>Resumen de Terapias por Paciente</h3>"
        html += f"<p><strong>Período:</strong> {fecha_inicio} a {fecha_fin}</p>"
        html += "<table border='1' style='width:100%; border-collapse:collapse; font-size:13px;'>"
        html += "<tr style='background:#004a80; color:white;'>"
        html += "<th style='padding:10px;'>Paciente</th>"
        html += "<th style='padding:10px;'>Total Citas</th>"
        html += "<th style='padding:10px;'>Servicios (Cantidad)</th>"
        html += "<th style='padding:10px;'>Profesionales</th>"
        html += "</tr>"
        for paciente, datos in sorted(resumen.items()):
            servicios_str = "; ".join([f"{s} ({c})" for s, c in datos["servicios"].items()])
            prof_str = ", ".join(sorted(datos["profesionales"]))
            html += f"<tr>"
            html += f"<td style='padding:8px;'>{paciente}</td>"
            html += f"<td style='padding:8px; text-align:center;'>{datos['citas']}</td>"
            html += f"<td style='padding:8px;'>{servicios_str}</td>"
            html += f"<td style='padding:8px;'>{prof_str}</td>"
            html += f"</tr>"
        html += "</table>"
        return jsonify({"html": html})
    
    # === REPORTE: Detallado ===
    elif tipo == "reporte_detallado":
        html = "<h3>📋 Reporte Detallado General (Todos los campos)</h3>"
        html += f"<p><strong>Período:</strong> {fecha_inicio} a {fecha_fin}</p>"
        html += "<table border='1' style='width:100%; border-collapse:collapse; font-size:12px;'>"
        html += "<tr style='background:#004a80; color:white;'>"
        html += "<th style='padding:8px;'>Tipo Doc</th><th style='padding:8px;'>Num Doc</th><th style='padding:8px;'>Paciente</th>"
        html += "<th style='padding:8px;'>Edad</th><th style='padding:8px;'>Género</th><th style='padding:8px;'>Régimen</th>"
        html += "<th style='padding:8px;'>CUPS</th><th style='padding:8px;'>Servicio</th><th style='padding:8px;'>Profesional</th>"
        html += "<th style='padding:8px;'>Fecha</th><th style='padding:8px;'>Hora</th><th style='padding:8px;'>Estado</th>"
        html += "</tr>"
        for r in registros:
            html += f"<tr>"
            html += f"<td style='padding:6px;'>{r.get('Tipo_Doc', '')}</td>"
            html += f"<td style='padding:6px;'>{r.get('Num_Doc', '')}</td>"
            html += f"<td style='padding:6px;'>{r.get('Nombre_Completo', '')}</td>"
            html += f"<td style='padding:6px;'>{r.get('Edad', '')}</td>"
            html += f"<td style='padding:6px;'>{r.get('Genero', '')}</td>"
            html += f"<td style='padding:6px;'>{r.get('REGIMEN', '')}</td>"
            html += f"<td style='padding:6px;'>{r.get('CUPS', '')}</td>"
            html += f"<td style='padding:6px;'>{r.get('Servicio', '')}</td>"
            html += f"<td style='padding:6px;'>{r.get('Profesional', '')}</td>"
            html += f"<td style='padding:6px;'>{r.get('Fecha', '')}</td>"
            html += f"<td style='padding:6px;'>{r.get('Hora', '')}</td>"
            html += f"<td style='padding:6px;'>{r.get('Estado', 'activa')}</td>"
            html += f"</tr>"
        html += "</table>"
        return jsonify({"html": html})
    
    # === ELSE FINAL: Tipo de reporte no válido ===
    else:
        return jsonify({"error": f"Tipo de reporte no válido: {tipo}"})

# === DASHBOARD ESTADÍSTICO ===
@app.route("/dashboard")
@login_requerido
def dashboard():
    from collections import defaultdict
    from datetime import datetime, date, timedelta
    rol = session.get('rol', 'admin')
    profesional_logueado = session.get('usuario', '')
    hoy = date.today()
    fecha_fin_str = request.args.get('fecha_fin', hoy.isoformat())
    fecha_inicio_str = request.args.get('fecha_inicio', (hoy - timedelta(days=90)).isoformat())
    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
        fecha_fin = datetime.strptime(fecha_fin_str, "%Y-%m-%d").date()
    except:
        fecha_inicio = hoy - timedelta(days=90)
        fecha_fin = hoy
    if rol == 'profesional':
        AGENDA = cargar_agenda_desde_csv_filtrada(profesional_logueado, solo_activas=False)
    else:
        AGENDA = cargar_agenda_desde_csv_filtrada(solo_activas=False)
    
    # === INICIALIZAR CONTADORES ===
    citas_por_mes = defaultdict(int)
    diagnosticos = defaultdict(int)
    cups_stats = defaultdict(int)
    estados = defaultdict(int)
    profesionales_stats = defaultdict(int)
    gestion_por_cups = defaultdict(lambda: {"subsidiado": 0, "contributivo": 0, "total": 0})
    
    total_citas = 0
    citas_activas = 0
    citas_canceladas = 0
    citas_no_asistio = 0
    
    for clave, citas in AGENDA.items():
        for cita in citas:
            try:
                fecha_cita = datetime.strptime(cita.get("Fecha", ""), "%Y-%m-%d").date()
                if fecha_inicio <= fecha_cita <= fecha_fin:
                    total_citas += 1
                    mes_key = fecha_cita.strftime("%Y-%m")
                    citas_por_mes[mes_key] += 1
                    dx = cita.get("Dx_Codigo", "SIN_DX")
                    diagnosticos[dx] += 1
                    cups = cita.get("CUPS", "SIN_CUPS")
                    servicio = cita.get("Servicio", cups)
                    cups_stats[f"{cups} - {servicio[:30]}"] += 1
                    estado = cita.get("Estado", "activa").strip().lower()
                    estados[estado] += 1
                    if estado == "activa":
                        citas_activas += 1
                    elif estado == "cancelada":
                        citas_canceladas += 1
                    elif estado == "no_asistio":
                        citas_no_asistio += 1
                    profesional = cita.get("Profesional", "Sin asignar")
                    profesionales_stats[profesional] += 1
                    
                    # ✅ CONTAR POR RÉGIMEN PARA GESTIÓN VS METAS
                    regimen = cita.get("REGIMEN", "").strip().upper()
                    if cups != "SIN_CUPS" and regimen in ("SUBSIDIADO", "CONTRIBUTIVO"):
                        if regimen == "SUBSIDIADO":
                            gestion_por_cups[cups]["subsidiado"] += 1
                        elif regimen == "CONTRIBUTIVO":
                            gestion_por_cups[cups]["contributivo"] += 1
                        gestion_por_cups[cups]["total"] += 1
            except Exception as e:
                print(f"Error procesando cita: {e}")
                continue
    
    # === CARGAR METAS DESDE ARCHIVO ===
    metas_cups = {}
    if os.path.exists("data/metas_mes.txt"):
        with open("data/metas_mes.txt", encoding="utf-8") as f:
            for linea in f:
                linea = linea.strip()
                if linea and "|" in linea and not linea.startswith("#"):
                    partes = linea.split("|")
                    if len(partes) >= 4:
                        cups_meta = partes[0].strip()
                        try:
                            metas_cups[cups_meta] = {
                                "subsidiado": int(partes[1].strip()),
                                "contributivo": int(partes[2].strip()),
                                "total": int(partes[3].strip())
                            }
                        except ValueError:
                            continue
    
    # === CALCULAR GESTIÓN VS METAS ===
    gestion_vs_metas = []
    for cups, gestion in sorted(gestion_por_cups.items(), key=lambda x: x[1]["total"], reverse=True):
        metas = metas_cups.get(cups, {"subsidiado": 0, "contributivo": 0, "total": 0})
        
        def calcular_pct(real, meta):
            return round((real / meta) * 100, 1) if meta > 0 else 0
        
        def color_pct(pct):
            return "green" if pct >= 100 else "orange" if pct >= 70 else "red"
        
        servicio_nombre = ""
        for clave, citas in AGENDA.items():
            for cita in citas:
                if cita.get("CUPS") == cups:
                    servicio_nombre = cita.get("Servicio", cups)[:25]
                    break
            if servicio_nombre:
                break
        
        gestion_vs_metas.append({
            "cups": cups,
            "servicio": servicio_nombre,
            "subsidiado": {
                "real": gestion["subsidiado"],
                "meta": metas["subsidiado"],
                "porcentaje": calcular_pct(gestion["subsidiado"], metas["subsidiado"]),
                "color": color_pct(calcular_pct(gestion["subsidiado"], metas["subsidiado"]))
            },
            "contributivo": {
                "real": gestion["contributivo"],
                "meta": metas["contributivo"],
                "porcentaje": calcular_pct(gestion["contributivo"], metas["contributivo"]),
                "color": color_pct(calcular_pct(gestion["contributivo"], metas["contributivo"]))
            },
            "total": {
                "real": gestion["total"],
                "meta": metas["total"],
                "porcentaje": calcular_pct(gestion["total"], metas["total"]),
                "color": color_pct(calcular_pct(gestion["total"], metas["total"]))
            }
        })
    
    # === PREPARAR DATOS PARA GRÁFICOS EXISTENTES ===
    meses_ordenados = sorted(citas_por_mes.keys())
    meses_labels = [datetime.strptime(m, "%Y-%m").strftime("%B %Y").capitalize() for m in meses_ordenados]
    meses_values = [citas_por_mes[m] for m in meses_ordenados]
    top_dx = sorted(diagnosticos.items(), key=lambda x: x[1], reverse=True)[:5]
    dx_labels = [f"{dx} ({cnt})" for dx, cnt in top_dx]
    dx_values = [cnt for _, cnt in top_dx]
    top_cups = sorted(cups_stats.items(), key=lambda x: x[1], reverse=True)[:5]
    cups_labels = [f"{cups[:20]}... ({cnt})" for cups, cnt in top_cups]
    cups_values = [cnt for _, cnt in top_cups]
    profesionales_ordenados = sorted(profesionales_stats.items(), key=lambda x: x[0].upper())
    prof_labels = [f"{prof} ({cnt})" for prof, cnt in profesionales_ordenados]
    prof_values = [cnt for _, cnt in profesionales_ordenados]
    estados_labels = list(estados.keys())
    estados_values = list(estados.values())
    
    print(f"=== DASHBOARD DEBUG ===")
    print(f"Total Citas: {total_citas}")
    print(f"Citas Activas: {citas_activas}")
    print(f"Citas Canceladas: {citas_canceladas}")
    print(f"Citas No Asistio: {citas_no_asistio}")
    print(f"======================")
    
    return render_template("dashboard.html",
        fecha_inicio=fecha_inicio_str,
        fecha_fin=fecha_fin_str,
        meses_labels=meses_labels,
        meses_values=meses_values,
        total_citas=total_citas,
        citas_activas=citas_activas,
        citas_canceladas=citas_canceladas,
        citas_no_asistio=citas_no_asistio,
        dx_labels=dx_labels,
        dx_values=dx_values,
        cups_labels=cups_labels,
        cups_values=cups_values,
        prof_labels=prof_labels,
        prof_values=prof_values,
        estados_labels=estados_labels,
        estados_values=estados_values,
        gestion_vs_metas=gestion_vs_metas,  # ✅ NUEVO: Enviar datos de gestión vs metas
        es_profesional=(rol == 'profesional'))

@app.route("/exportar-gestion-metas", methods=["POST"])
@login_requerido
def exportar_gestion_metas():
    """Exporta datos de Gestión vs Metas por CUPS y Régimen a Excel"""
    from collections import defaultdict
    from datetime import datetime, date, timedelta
    import pandas as pd
    from io import BytesIO
    
    data = request.get_json()
    fecha_inicio_str = data.get("fecha_inicio", (date.today() - timedelta(days=90)).isoformat())
    fecha_fin_str = data.get("fecha_fin", date.today().isoformat())
    
    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
        fecha_fin = datetime.strptime(fecha_fin_str, "%Y-%m-%d").date()
    except:
        return jsonify({"error": "Formato de fecha inválido."}), 400
    
    # Cargar agenda
    rol = session.get('rol', 'admin')
    profesional_logueado = session.get('usuario', '')
    if rol == 'profesional':
        AGENDA = cargar_agenda_desde_csv_filtrada(profesional_logueado, solo_activas=False)
    else:
        AGENDA = cargar_agenda_desde_csv_filtrada(solo_activas=False)
    
    # Contar citas por CUPS y régimen
    gestion_por_cups = defaultdict(lambda: {"subsidiado": 0, "contributivo": 0, "total": 0})
    for clave, citas in AGENDA.items():
        for cita in citas:
            try:
                fecha_cita = datetime.strptime(cita.get("Fecha", ""), "%Y-%m-%d").date()
                if fecha_inicio <= fecha_cita <= fecha_fin:
                    cups = cita.get("CUPS", "").strip()
                    regimen = cita.get("REGIMEN", "").strip().upper()
                    servicio = cita.get("Servicio", cups)
                    if cups and cups != "SIN_CUPS" and regimen in ("SUBSIDIADO", "CONTRIBUTIVO"):
                        if regimen == "SUBSIDIADO":
                            gestion_por_cups[cups]["subsidiado"] += 1
                        elif regimen == "CONTRIBUTIVO":
                            gestion_por_cups[cups]["contributivo"] += 1
                        gestion_por_cups[cups]["total"] += 1
                        gestion_por_cups[cups]["servicio"] = servicio
            except:
                continue
    
    # Cargar metas
    metas_cups = {}
    if os.path.exists("data/metas_mes.txt"):
        with open("data/metas_mes.txt", encoding="utf-8") as f:
            for linea in f:
                linea = linea.strip()
                if linea and "|" in linea and not linea.startswith("#"):
                    partes = linea.split("|")
                    if len(partes) >= 4:
                        cups_meta = partes[0].strip()
                        try:
                            metas_cups[cups_meta] = {
                                "subsidiado": int(partes[1].strip()),
                                "contributivo": int(partes[2].strip()),
                                "total": int(partes[3].strip())
                            }
                        except ValueError:
                            continue
    
    # Preparar datos para Excel
    filas = []
    for cups, datos in sorted(gestion_por_cups.items(), key=lambda x: x[1]["total"], reverse=True):
        metas = metas_cups.get(cups, {"subsidiado": 0, "contributivo": 0, "total": 0})
        servicio = datos.get("servicio", cups)
        
        def calcular_pct(real, meta):
            return round((real / meta) * 100, 1) if meta > 0 else None
        
        filas.append({
            "CUPS": cups,
            "Servicio": servicio,
            # Subsidiado
            "Real Subsidiado": datos["subsidiado"],
            "Meta Subsidiado": metas["subsidiado"],
            "Diferencia Subsidiado": datos["subsidiado"] - metas["subsidiado"],
            "% Cumplimiento Subsidiado": calcular_pct(datos["subsidiado"], metas["subsidiado"]),
            # Contributivo
            "Real Contributivo": datos["contributivo"],
            "Meta Contributivo": metas["contributivo"],
            "Diferencia Contributivo": datos["contributivo"] - metas["contributivo"],
            "% Cumplimiento Contributivo": calcular_pct(datos["contributivo"], metas["contributivo"]),
            # Total
            "Real Total": datos["total"],
            "Meta Total": metas["total"],
            "Diferencia Total": datos["total"] - metas["total"],
            "% Cumplimiento Total": calcular_pct(datos["total"], metas["total"]),
        })
    
    # Crear DataFrame y exportar
    df = pd.DataFrame(filas)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Gestion_vs_Metas")
        # Agregar formato condicional básico (opcional)
        worksheet = writer.sheets["Gestion_vs_Metas"]
        worksheet.column_dimensions['A'].width = 12  # CUPS
        worksheet.column_dimensions['B'].width = 40  # Servicio
        for col in ['D', 'G', 'J']:  # Columnas de % Cumplimiento
            if col in worksheet.column_dimensions:
                worksheet.column_dimensions[col].width = 18
    
    output.seek(0)
    nombre_archivo = f"gestion_metas_{fecha_inicio_str}_a_{fecha_fin_str}.xlsx"
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre_archivo
    )
@app.route("/exportar-control-terapias", methods=["POST"])
@login_requerido
def exportar_control_terapias():
    """Exporta reporte Control de Terapias a Excel"""
    data = request.get_json()
    fecha_inicio = data.get("fecha_inicio")
    fecha_fin = data.get("fecha_fin")
    
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"error": "No hay datos para generar el reporte."}), 400
    
    try:
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        if fi > ff:
            return jsonify({"error": "La fecha de inicio no puede ser mayor que la final."}), 400
    except:
        return jsonify({"error": "Formato de fecha inválido."}), 400
    
    control = defaultdict(lambda: {
        "autorizadas": 0, 
        "realizadas": 0,
        "no_asistio": 0,
        "nombre": "", 
        "servicio": "", 
        "documento": "",
        "profesional": "",
        "fecha_registro": ""
    })
    
    # === PASO 1: Contar AUTORIZADAS (TODAS las filas NO canceladas) ===
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            estado = row.get("Estado", "activa").strip().lower()
            
            if estado == "cancelada":
                continue
            
            profesional = row.get("Profesional", "Sin asignar")
            cups = row.get("CUPS", "SIN_CUPS")
            clave = (row.get("Tipo_Doc"), row.get("Num_Doc"), profesional, cups)
            
            control[clave]["nombre"] = row.get("Nombre_Completo", "Desconocido")
            control[clave]["servicio"] = f"{cups} - {row.get('Servicio', '')}"
            control[clave]["documento"] = f"{row.get('Tipo_Doc', '')}-{row.get('Num_Doc', '')}"
            control[clave]["profesional"] = profesional
            control[clave]["fecha_registro"] = row.get("Fecha_Registro", "")[:10]
            control[clave]["autorizadas"] += 1
    
    # === PASO 2: Contar REALIZADAS y NO ASISTIO (dentro del rango) ===
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                fecha_cita = datetime.strptime(row["Fecha"], "%Y-%m-%d").date()
                if fi <= fecha_cita <= ff:
                    estado = row.get("Estado", "activa").strip().lower()
                    
                    if estado == "cancelada":
                        continue
                    
                    profesional = row.get("Profesional", "Sin asignar")
                    cups = row.get("CUPS", "SIN_CUPS")
                    clave = (row.get("Tipo_Doc"), row.get("Num_Doc"), profesional, cups)
                    
                    if estado == "activa":
                        control[clave]["realizadas"] += 1
                    
                    if estado == "no_asistio":
                        control[clave]["no_asistio"] += 1
            except:
                continue
    
    # ✅ PREPARAR DATOS PARA EXCEL
    filas = []
    for clave, datos in sorted(control.items(), key=lambda x: (x[1]["nombre"], x[1]["profesional"])):
        aut = datos["autorizadas"]
        real = datos["realizadas"]
        nas = datos["no_asistio"]
        pend = max(0, aut - real - nas)
        pct = round((real / aut * 100), 1) if aut > 0 else 0
        
        # ✅ FILTRO: NO incluir órdenes donde todas están canceladas
        if aut == 0:
            continue
        
        filas.append({
            "Documento": datos["documento"],
            "Paciente": datos["nombre"],
            "Profesional": datos["profesional"],
            "Servicio (CUPS)": datos["servicio"],
            "Fecha Orden": datos["fecha_registro"],
            "Autorizadas": aut,
            "Realizadas": real,
            "No Asistió": nas,
            "Pendientes": pend,
            "% Cumplimiento": pct
        })
    
    # ✅ CREAR EXCEL
    df = pd.DataFrame(filas)
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Control_Terapias")
        
        worksheet = writer.sheets["Control_Terapias"]
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 35
        worksheet.column_dimensions['C'].width = 25
        worksheet.column_dimensions['D'].width = 40
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 12
        worksheet.column_dimensions['G'].width = 12
        worksheet.column_dimensions['H'].width = 12
        worksheet.column_dimensions['I'].width = 12
        worksheet.column_dimensions['J'].width = 18
        
        # Formato condicional para % Cumplimiento
        from openpyxl.styles import PatternFill, Font
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=10, max_col=10):
            for cell in row:
                if cell.value is not None:
                    try:
                        pct_val = float(cell.value)
                        if pct_val >= 100:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100", bold=True)
                        elif pct_val >= 70:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            cell.font = Font(color="9C5700", bold=True)
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006", bold=True)
                    except:
                        pass
    
    output.seek(0)
    nombre_archivo = f"control_terapias_{fecha_inicio}_a_{fecha_fin}.xlsx"
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre_archivo
    )

@app.route("/exportar-actividad-profesional", methods=["POST"])
@login_requerido
def exportar_actividad_profesional():
    """Exporta reporte de Actividad Profesional consolidado a Excel"""
    data = request.get_json()
    fecha_inicio = data.get("fecha_inicio")
    fecha_fin = data.get("fecha_fin")
    
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"error": "No hay datos para generar el reporte."}), 400
    
    try:
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        if fi > ff:
            return jsonify({"error": "La fecha de inicio no puede ser mayor que la final."}), 400
    except:
        return jsonify({"error": "Formato de fecha inválido."}), 400
    
    # === CONTAR CITAS POR PROFESIONAL Y SERVICIO ===
    resumen = defaultdict(lambda: defaultdict(int))
    profesionales_pacientes = defaultdict(set)
    
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                fecha_cita = datetime.strptime(row["Fecha"], "%Y-%m-%d").date()
                if fi <= fecha_cita <= ff:
                    # ✅ SOLO CITAS ACTIVAS (excluir canceladas y no_asistio)
                    if row.get("Estado", "activa").strip().lower() == "activa":
                        prof = row["Profesional"]
                        servicio = f"{row['Servicio']} ({row['CUPS']})"
                        resumen[prof][servicio] += 1
                        profesionales_pacientes[prof].add(row["Num_Doc"])
            except:
                continue
    
    # === PREPARAR DATOS PARA EXCEL ===
    filas = []
    for prof in sorted(resumen.keys()):
        servicios = resumen[prof]
        total_pacientes = len(profesionales_pacientes[prof])
        total_citas = sum(servicios.values())
        
        for servicio, cant in sorted(servicios.items()):
            filas.append({
                "Profesional": prof,
                "Servicio (CUPS)": servicio,
                "Citas": cant,
                "Pacientes Atendidos": total_pacientes,
                "Total Citas Profesional": total_citas
            })
    
    # === CREAR EXCEL ===
    df = pd.DataFrame(filas)
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Actividad_Profesional")
        
        worksheet = writer.sheets["Actividad_Profesional"]
        
        # Ajustar anchos de columna
        worksheet.column_dimensions['A'].width = 35  # Profesional
        worksheet.column_dimensions['B'].width = 45  # Servicio
        worksheet.column_dimensions['C'].width = 12  # Citas
        worksheet.column_dimensions['D'].width = 18  # Pacientes
        worksheet.column_dimensions['E'].width = 20  # Total
    
    output.seek(0)
    nombre_archivo = f"actividad_profesional_{fecha_inicio}_a_{fecha_fin}.xlsx"
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre_archivo
    )
@app.route("/exportar-terapias-paciente", methods=["POST"])
@login_requerido
def exportar_terapias_paciente():
    """Exporta reporte de Terapias por Paciente consolidado a Excel"""
    data = request.get_json()
    fecha_inicio = data.get("fecha_inicio")
    fecha_fin = data.get("fecha_fin")
    
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"error": "No hay datos para generar el reporte."}), 400
    
    try:
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        if fi > ff:
            return jsonify({"error": "La fecha de inicio no puede ser mayor que la final."}), 400
    except:
        return jsonify({"error": "Formato de fecha inválido."}), 400
    
    # === CONTAR TERAPIAS POR PACIENTE (SOLO ACTIVAS) ===
    resumen = defaultdict(lambda: {"citas": 0, "servicios": defaultdict(int), "profesionales": set()})
    
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                fecha_cita = datetime.strptime(row["Fecha"], "%Y-%m-%d").date()
                if fi <= fecha_cita <= ff:
                    # ✅ SOLO CITAS ACTIVAS (excluir canceladas y no_asistio)
                    if row.get("Estado", "activa").strip().lower() == "activa":
                        paciente = f"{row['Nombre_Completo']} ({row['Tipo_Doc']}-{row['Num_Doc']})"
                        resumen[paciente]["citas"] += 1
                        resumen[paciente]["servicios"][row["Servicio"]] += 1
                        resumen[paciente]["profesionales"].add(row["Profesional"])
            except:
                continue
    
    # === PREPARAR DATOS PARA EXCEL ===
    filas = []
    for paciente, datos in sorted(resumen.items()):
        servicios_str = "; ".join([f"{s} ({c})" for s, c in datos["servicios"].items()])
        prof_str = ", ".join(sorted(datos["profesionales"]))
        
        filas.append({
            "Paciente": paciente,
            "Total Citas": datos["citas"],
            "Servicios (Cantidad)": servicios_str,
            "Profesionales": prof_str
        })
    
    # === CREAR EXCEL ===
    df = pd.DataFrame(filas)
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Terapias_Paciente")
        
        worksheet = writer.sheets["Terapias_Paciente"]
        
        # Ajustar anchos de columna
        worksheet.column_dimensions['A'].width = 50  # Paciente
        worksheet.column_dimensions['B'].width = 12  # Total Citas
        worksheet.column_dimensions['C'].width = 60  # Servicios
        worksheet.column_dimensions['D'].width = 40  # Profesionales
    
    output.seek(0)
    nombre_archivo = f"terapias_paciente_{fecha_inicio}_a_{fecha_fin}.xlsx"
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre_archivo
    )

@app.route("/exportar-reporte", methods=["POST"])
@login_requerido
def exportar_reporte():
    tipo = request.form.get("tipo")
    fecha_inicio = request.form.get("fecha_inicio")
    fecha_fin = request.form.get("fecha_fin")
    
    if not os.path.exists("data/agenda.csv"):
        return "No hay datos para generar reportes.", 400
    
    try:
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        if fi > ff:
            return "La fecha de inicio no puede ser mayor que la final.", 400
    except:
        return "Formato de fecha inválido.", 400
    
    registros = []
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                fecha_cita = datetime.strptime(row["Fecha"], "%Y-%m-%d").date()
                if fi <= fecha_cita <= ff:
                    registros.append(row)
            except:
                continue
    
    if not registros:
        return "No hay registros en el rango de fechas seleccionado.", 400
    
    rol = session.get('rol', 'admin')
    profesional_logueado = session.get('usuario', '')
    if rol == 'profesional':
        registros = [r for r in registros if r.get("Profesional") == profesional_logueado]
    
    # === REPORTE: Actividad Profesional (CONSOLIDADO) ===
    if tipo == "actividad_profesional":
        resumen = defaultdict(lambda: defaultdict(int))
        profesionales_pacientes = defaultdict(set)
        for r in registros:
            if r.get("Estado", "activa") == "activa":
                prof = r["Profesional"]
                servicio = f"{r['Servicio']} ({r['CUPS']})"
                resumen[prof][servicio] += 1
                profesionales_pacientes[prof].add(r["Num_Doc"])
        
        filas = []
        for prof in sorted(resumen.keys()):
            servicios = resumen[prof]
            total_pacientes = len(profesionales_pacientes[prof])
            for servicio, cant in sorted(servicios.items()):
                filas.append({
                    "Profesional": prof,
                    "Servicio (CUPS)": servicio,
                    "Citas": cant,
                    "Pacientes Atendidos": total_pacientes
                })
        
        df = pd.DataFrame(filas)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Actividad_Profesional")
            worksheet = writer.sheets["Actividad_Profesional"]
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 45
            worksheet.column_dimensions['C'].width = 12
            worksheet.column_dimensions['D'].width = 18
        output.seek(0)
        nombre_archivo = f"actividad_profesional_{fecha_inicio}_a_{fecha_fin}.xlsx"
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=nombre_archivo)
    
    # === REPORTE: Terapias por Paciente (CONSOLIDADO) ===
    elif tipo == "terapias_paciente":
        resumen = defaultdict(lambda: {"citas": 0, "servicios": defaultdict(int), "profesionales": set()})
        for r in registros:
            if r.get("Estado", "activa") == "activa":
                paciente = f"{r['Nombre_Completo']} ({r['Tipo_Doc']}-{r['Num_Doc']})"
                resumen[paciente]["citas"] += 1
                resumen[paciente]["servicios"][r["Servicio"]] += 1
                resumen[paciente]["profesionales"].add(r["Profesional"])
        
        filas = []
        for paciente, datos in sorted(resumen.items()):
            servicios_str = "; ".join([f"{s} ({c})" for s, c in datos["servicios"].items()])
            prof_str = ", ".join(sorted(datos["profesionales"]))
            filas.append({
                "Paciente": paciente,
                "Total Citas": datos["citas"],
                "Servicios (Cantidad)": servicios_str,
                "Profesionales": prof_str
            })
        
        df = pd.DataFrame(filas)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Terapias_Paciente")
            worksheet = writer.sheets["Terapias_Paciente"]
            worksheet.column_dimensions['A'].width = 50
            worksheet.column_dimensions['B'].width = 12
            worksheet.column_dimensions['C'].width = 60
            worksheet.column_dimensions['D'].width = 40
        output.seek(0)
        nombre_archivo = f"terapias_paciente_{fecha_inicio}_a_{fecha_fin}.xlsx"
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=nombre_archivo)
    
    # === REPORTE: Control de Terapias (CONSOLIDADO) ===
    elif tipo == "control_terapias":
        control = defaultdict(lambda: {
            "autorizadas": 0, "realizadas": 0, "no_asistio": 0,
            "nombre": "", "servicio": "", "documento": "",
            "profesional": "", "fecha_registro": ""
        })
        
        # PASO 1: Contar AUTORIZADAS (todas las filas NO canceladas en TODO el CSV)
        with open("data/agenda.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                estado = row.get("Estado", "activa").strip().lower()
                if estado == "cancelada":
                    continue
                profesional = row.get("Profesional", "Sin asignar")
                cups = row.get("CUPS", "SIN_CUPS")
                clave = (row.get("Tipo_Doc"), row.get("Num_Doc"), profesional, cups)
                control[clave]["nombre"] = row.get("Nombre_Completo", "Desconocido")
                control[clave]["servicio"] = f"{cups} - {row.get('Servicio', '')}"
                control[clave]["documento"] = f"{row.get('Tipo_Doc', '')}-{row.get('Num_Doc', '')}"
                control[clave]["profesional"] = profesional
                control[clave]["fecha_registro"] = row.get("Fecha_Registro", "")[:10]
                control[clave]["autorizadas"] += 1
        
        # PASO 2: Contar REALIZADAS y NO ASISTIO (solo dentro del rango)
        for r in registros:
            estado = r.get("Estado", "activa").strip().lower()
            if estado == "cancelada":
                continue
            profesional = r.get("Profesional", "Sin asignar")
            cups = r.get("CUPS", "SIN_CUPS")
            clave = (r.get("Tipo_Doc"), r.get("Num_Doc"), profesional, cups)
            if estado == "activa":
                control[clave]["realizadas"] += 1
            if estado == "no_asistio":
                control[clave]["no_asistio"] += 1
        
        filas = []
        for clave, datos in sorted(control.items(), key=lambda x: (x[1]["nombre"], x[1]["profesional"])):
            aut = datos["autorizadas"]
            real = datos["realizadas"]
            nas = datos["no_asistio"]
            pend = max(0, aut - real - nas)
            pct = round((real / aut * 100), 1) if aut > 0 else 0
            if aut == 0:
                continue
            filas.append({
                "Documento": datos["documento"],
                "Paciente": datos["nombre"],
                "Profesional": datos["profesional"],
                "Servicio (CUPS)": datos["servicio"],
                "Fecha Orden": datos["fecha_registro"],
                "Autorizadas": aut,
                "Realizadas": real,
                "No Asistió": nas,
                "Pendientes": pend,
                "% Cumplimiento": pct
            })
        
        df = pd.DataFrame(filas)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Control_Terapias")
            worksheet = writer.sheets["Control_Terapias"]
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 35
            worksheet.column_dimensions['C'].width = 25
            worksheet.column_dimensions['D'].width = 40
            worksheet.column_dimensions['E'].width = 15
            worksheet.column_dimensions['F'].width = 12
            worksheet.column_dimensions['G'].width = 12
            worksheet.column_dimensions['H'].width = 12
            worksheet.column_dimensions['I'].width = 12
            worksheet.column_dimensions['J'].width = 18
        output.seek(0)
        nombre_archivo = f"control_terapias_{fecha_inicio}_a_{fecha_fin}.xlsx"
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=nombre_archivo)
    
    # === REPORTE: Diagnósticos (CONSOLIDADO) ===
    elif tipo == "reporte_diagnosticos":
        registros_activos = [r for r in registros if r.get("Estado", "activa") == "activa"]
        diag_data = defaultdict(lambda: {"citas": 0, "pacientes": set()})
        for r in registros_activos:
            dx_codigo = r.get("Dx_Codigo", "SIN_DX")
            dx_desc = r.get("Dx_Descripcion", "No especificado")
            clave = f"{dx_codigo} - {dx_desc}"
            diag_data[clave]["citas"] += 1
            diag_data[clave]["pacientes"].add(f"{r['Tipo_Doc']}-{r['Num_Doc']}")
        
        filas = []
        for dx_completo, datos in sorted(diag_data.items(), key=lambda x: x[1]['citas'], reverse=True):
            partes = dx_completo.split(" - ", 1)
            codigo_dx = partes[0]
            desc_dx = partes[1] if len(partes) > 1 else "Sin descripción"
            filas.append({
                "Código Dx": codigo_dx,
                "Diagnóstico": desc_dx,
                "Total Citas": datos["citas"],
                "Pacientes Únicos": len(datos["pacientes"])
            })
        
        df = pd.DataFrame(filas)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Diagnosticos")
            worksheet = writer.sheets["Diagnosticos"]
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 50
            worksheet.column_dimensions['C'].width = 12
            worksheet.column_dimensions['D'].width = 18
        output.seek(0)
        nombre_archivo = f"diagnosticos_{fecha_inicio}_a_{fecha_fin}.xlsx"
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=nombre_archivo)
    
    # === REPORTE: General (CONSOLIDADO) ===
    elif tipo == "reporte_general":
        total_citas = len([r for r in registros if r.get("Estado", "activa") == "activa"])
        pacientes_unicos = len(set(r["Num_Doc"] for r in registros if r.get("Estado", "activa") == "activa"))
        profesionales_unicos = len(set(r["Profesional"] for r in registros if r.get("Estado", "activa") == "activa"))
        servicios = defaultdict(int)
        for r in registros:
            servicios[r["CUPS"]] += 1
        
        filas = []
        for cups, count in sorted(servicios.items(), key=lambda x: x[1], reverse=True):
            filas.append({
                "CUPS": cups,
                "Cantidad": count
            })
        
        # Hoja 1: Resumen
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_resumen = pd.DataFrame([{
                "Métrica": ["Total Citas Activas", "Pacientes Atendidos", "Profesionales Activos"],
                "Valor": [total_citas, pacientes_unicos, profesionales_unicos]
            }])
            df_resumen.to_excel(writer, index=False, sheet_name="Resumen")
            
            # Hoja 2: Servicios
            df_servicios = pd.DataFrame(filas)
            df_servicios.to_excel(writer, index=False, sheet_name="Servicios")
            
            worksheet = writer.sheets["Resumen"]
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 15
            
            worksheet2 = writer.sheets["Servicios"]
            worksheet2.column_dimensions['A'].width = 15
            worksheet2.column_dimensions['B'].width = 12
        
        output.seek(0)
        nombre_archivo = f"reporte_general_{fecha_inicio}_a_{fecha_fin}.xlsx"
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=nombre_archivo)
    
    # === REPORTE: Por Paciente (CONSOLIDADO) ===
    elif tipo == "reporte_por_paciente":
        tipo_doc_filtro = request.form.get("tipo_doc_paciente", "").strip().upper()
        num_doc_filtro = request.form.get("num_doc_paciente", "").strip()
        
        if not tipo_doc_filtro or not num_doc_filtro:
            return "Debe ingresar tipo y número de documento.", 400
        
        citas_paciente = [r for r in registros if r.get("Tipo_Doc") == tipo_doc_filtro and r.get("Num_Doc") == num_doc_filtro]
        
        if not citas_paciente:
            return f"No se encontraron citas para {tipo_doc_filtro}-{num_doc_filtro}.", 400
        
        citas_paciente.sort(key=lambda x: x.get("Fecha", ""), reverse=True)
        
        filas = []
        for r in citas_paciente:
            estado = r.get("Estado", "activa")
            estado_label = "Programada" if estado == "activa" else "Cancelada" if estado == "cancelada" else "No Asistió" if estado == "no_asistio" else estado
            filas.append({
                "Fecha": r.get("Fecha", ""),
                "Hora": r.get("Hora", ""),
                "Profesional": r.get("Profesional", ""),
                "Servicio (CUPS)": f"{r.get('Servicio', '')} ({r.get('CUPS', '')})",
                "Estado": estado_label,
                "Dx": r.get("Dx_Codigo", ""),
                "Teléfono": r.get("Celular", "")
            })
        
        df = pd.DataFrame(filas)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Citas_Paciente")
            worksheet = writer.sheets["Citas_Paciente"]
            worksheet.column_dimensions['A'].width = 12
            worksheet.column_dimensions['B'].width = 10
            worksheet.column_dimensions['C'].width = 35
            worksheet.column_dimensions['D'].width = 40
            worksheet.column_dimensions['E'].width = 15
            worksheet.column_dimensions['F'].width = 15
            worksheet.column_dimensions['G'].width = 15
        output.seek(0)
        nombre_archivo = f"citas_paciente_{tipo_doc_filtro}_{num_doc_filtro}_{fecha_inicio}_a_{fecha_fin}.xlsx"
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=nombre_archivo)
    
    # === REPORTE: Canceladas (DATOS CRUDOS - FILTRADO) ===
    elif tipo == "reporte_canceladas":
        canceladas = [r for r in registros if r.get("Estado") == "cancelada"]
        if not canceladas:
            return "No hay citas canceladas en el rango seleccionado.", 400
        
        columnas = ["Fecha", "Hora", "Nombre_Completo", "Tipo_Doc", "Num_Doc", "Servicio", "CUPS", "Profesional", "Dx_Codigo", "Celular", "Fecha_Registro"]
        df = pd.DataFrame([{k: r.get(k, "") for k in columnas} for r in canceladas])
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Canceladas")
        output.seek(0)
        nombre_archivo = f"canceladas_{fecha_inicio}_a_{fecha_fin}.xlsx"
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=nombre_archivo)
    
    # === REPORTE: No Asistieron (DATOS CRUDOS - FILTRADO) ===
    elif tipo == "reporte_no_asistieron":
        no_asistieron = [r for r in registros if r.get("Estado") == "no_asistio"]
        if not no_asistieron:
            return "No hay registros de inasistencia en el rango seleccionado.", 400
        
        columnas = ["Fecha", "Hora", "Nombre_Completo", "Tipo_Doc", "Num_Doc", "Servicio", "CUPS", "Profesional", "Dx_Codigo", "Celular", "Fecha_Registro"]
        df = pd.DataFrame([{k: r.get(k, "") for k in columnas} for r in no_asistieron])
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="No_Asistieron")
        output.seek(0)
        nombre_archivo = f"no_asistieron_{fecha_inicio}_a_{fecha_fin}.xlsx"
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=nombre_archivo)
    
    # === REPORTE: Detallado (DATOS CRUDOS - TODAS LAS COLUMNAS) ===
    elif tipo == "reporte_detallado":
        columnas = [
            "Tipo_Doc", "Num_Doc", "Nombre_Completo", "Edad", "Genero",
            "REGIMEN", "TIPO_AFILIADO", "CODIGO_EPS", "DEPARTAMENTO", "CIUDAD", "ZONA", "ESTADO",
            "Celular", "Email", "Dx_Codigo", "Dx_Descripcion",
            "CUPS", "Servicio", "Profesional", "Fecha", "Hora",
            "Cantidad_Total", "Frecuencia_Semanal", "Duracion_Meses", "Observacion", "Estado"
        ]
        df = pd.DataFrame([{k: r.get(k, "") for k in columnas} for r in registros])
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Reporte_Detallado")
        output.seek(0)
        nombre_archivo = f"reporte_detallado_{fecha_inicio}_a_{fecha_fin}.xlsx"
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=nombre_archivo)
    
    # === ELSE: Tipo no válido ===
    else:
        return "Tipo de reporte no válido.", 400
@app.route("/reasignar-profesional", methods=["POST"])
@login_requerido
def reasignar_profesional():
    if session.get('rol') != 'admin':
        return jsonify({"exito": False, "mensaje": "Acceso denegado."})
    
    data = request.get_json()
    profesional_saliente = data.get("profesional_saliente", "").strip()
    profesional_nuevo = data.get("profesional_nuevo", "").strip()
    
    if not profesional_saliente or not profesional_nuevo:
        return jsonify({"exito": False, "mensaje": "Faltan datos."})
    
    if profesional_saliente == profesional_nuevo:
        return jsonify({"exito": False, "mensaje": "El profesional de origen y destino son iguales."})
    
    profesionales_activos = set(cargar_profesionales("data/Profesionales.txt"))
    if profesional_nuevo not in profesionales_activos:
        return jsonify({"exito": False, "mensaje": "El profesional destino no está activo."})
    
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"exito": False, "mensaje": "No hay agenda registrada."})
    
    # ✅ NUEVO: Obtener fecha de hoy para filtro
    fecha_hoy = date.today().isoformat()
    
    agenda_actual = cargar_agenda_desde_csv_filtrada()
    registros_actualizados = []
    reasignadas = 0
    omitidas = 0
    citas_pasadas_preservadas = 0  # ✅ NUEVO: Contador para informar
    
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or CAMPOS_AGENDA
        
        for row in reader:
            # ✅ CORREGIDO: TODAS las filas se agregan (evita borrado masivo)
            # ✅ NUEVO: Solo reasignar si fecha es HOY o FUTURA
            if (row.get("Profesional") == profesional_saliente and
                row.get("Estado", "activa") == "activa"):
                
                fecha_cita = row.get("Fecha", "")
                
                # ✅ FILTRO CRÍTICO: ¿La cita es pasada?
                if fecha_cita < fecha_hoy:
                    citas_pasadas_preservadas += 1
                    registros_actualizados.append(row)
                    continue
                
                # ✅ Solo procesar citas futuras
                hora = row["Hora"]
                clave_nueva = f"{profesional_nuevo}|{fecha_cita}|{hora}"
                
                if len(agenda_actual.get(clave_nueva, [])) < 2:
                    row["Profesional"] = profesional_nuevo
                    reasignadas += 1
                    if clave_nueva not in agenda_actual:
                        agenda_actual[clave_nueva] = []
                    agenda_actual[clave_nueva].append(row)
                else:
                    omitidas += 1
            
            # ✅ CRÍTICO: TODAS las filas se agregan (otros profesionales, canceladas, etc.)
            registros_actualizados.append(row)
    
    with open("data/agenda.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(registros_actualizados)
    
    # ✅ MENSAJE MEJORADO: Informa cuántas citas pasadas se preservaron
    mensaje = f"✅ {reasignadas} citas FUTURAS reasignadas a '{profesional_nuevo}'."
    if citas_pasadas_preservadas > 0:
        mensaje += f" 📁 {citas_pasadas_preservadas} citas PASADAS preservadas con '{profesional_saliente}'."
    if omitidas > 0:
        mensaje += f" ⚠️ {omitidas} citas no pudieron reasignarse (horario lleno)."
    
    return jsonify({"exito": True, "mensaje": mensaje})
@app.route("/reasignar-cita", methods=["POST"])
@login_requerido
def reasignar_cita():
    if session.get('rol') != 'admin':
        return jsonify({"exito": False, "mensaje": "Acceso denegado."})
    data = request.get_json()
    required = ["tipo_doc", "num_doc", "profesional_origen", "fecha_origen", "hora_origen", "cups", "nuevo_profesional", "nueva_fecha", "nueva_hora"]
    if not all(k in data for k in required):
        return jsonify({"exito": False, "mensaje": "Faltan datos para reasignar la cita."})
    profesionales_activos = set(cargar_profesionales("data/Profesionales.txt"))
    if data["nuevo_profesional"] not in profesionales_activos:
        return jsonify({"exito": False, "mensaje": "El nuevo profesional no está activo."})
    if not es_dia_habil(data["nueva_fecha"]):
        return jsonify({"exito": False, "mensaje": "La nueva fecha no es hábil."})
    agenda_actual = cargar_agenda_desde_csv_filtrada()
    clave_nueva = f"{data['nuevo_profesional']}|{data['nueva_fecha']}|{data['nueva_hora']}"
    if len(agenda_actual.get(clave_nueva, [])) >= 2:
        return jsonify({"exito": False, "mensaje": "El nuevo horario ya está completo (máx. 2 citas)."})
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"exito": False, "mensaje": "No hay agenda registrada."})
    registros_actualizados = []
    cita_encontrada = False
    fieldnames = None
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or CAMPOS_AGENDA
        for row in reader:
            if (row.get("Tipo_Doc") == data["tipo_doc"] and
                row.get("Num_Doc") == data["num_doc"] and
                row.get("Profesional") == data["profesional_origen"] and
                row.get("Fecha") == data["fecha_origen"] and
                row.get("Hora") == data["hora_origen"] and
                row.get("CUPS") == data["cups"] and
                row.get("Estado", "activa") == "activa"):
                row["Profesional"] = data["nuevo_profesional"]
                row["Fecha"] = data["nueva_fecha"]
                row["Hora"] = data["nueva_hora"]
                cita_encontrada = True
            registros_actualizados.append(row)
    if not cita_encontrada:
        return jsonify({"exito": False, "mensaje": "Cita original no encontrada o ya cancelada."})
    with open("data/agenda.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(registros_actualizados)
    return jsonify({"exito": True, "mensaje": "✅ Cita reasignada correctamente."})
@app.route("/reasignar-bloque", methods=["POST"])
@login_requerido
def reasignar_bloque():
    if session.get('rol') != 'admin':
        return jsonify({"exito": False, "mensaje": "Acceso denegado."})
    
    data = request.get_json()
    required = ["tipo_doc", "num_doc", "cups", "profesional_origen", "nuevo_profesional"]
    if not all(k in data for k in required):
        return jsonify({"exito": False, "mensaje": "Faltan datos para reasignar el bloque."})
    
    profesionales_activos = set(cargar_profesionales("data/Profesionales.txt"))
    if data["nuevo_profesional"] not in profesionales_activos:
        return jsonify({"exito": False, "mensaje": "El nuevo profesional no está activo."})
    
    if data["profesional_origen"] == data["nuevo_profesional"]:
        return jsonify({"exito": False, "mensaje": "El profesional de origen y destino son iguales."})
    
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"exito": False, "mensaje": "No hay agenda registrada."})
    
    # ✅ NUEVO: Obtener fecha de hoy para filtro
    fecha_hoy = date.today().isoformat()
    
    agenda_actual = cargar_agenda_desde_csv_filtrada()
    registros_actualizados = []
    citas_reasignadas = 0
    citas_pasadas_preservadas = 0  # ✅ NUEVO: Contador
    conflictos = []
    block_rows = []
    
    # 🔴 PRIMERA LECTURA: Identificar filas del bloque
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or CAMPOS_AGENDA
        
        for row in reader:
            if (row.get("Tipo_Doc") == data["tipo_doc"] and
                row.get("Num_Doc") == data["num_doc"] and
                row.get("CUPS") == data["cups"] and
                row.get("Profesional") == data["profesional_origen"] and
                row.get("Estado", "activa") == "activa"):
                block_rows.append(row)
    
    if not block_rows:
        return jsonify({"exito": False, "mensaje": "No se encontró un bloque activo con esos datos."})
    
    # ✅ Validar conflictos SOLO para citas futuras
    for row in block_rows:
        fecha = row['Fecha']
        
        # ✅ Si es cita pasada, no validar conflictos (se preserva)
        if fecha < fecha_hoy:
            continue
        
        hora = row['Hora']
        dx_paciente = row.get("Dx_Codigo", "").strip().upper()
        cups_paciente = row.get("CUPS", "").strip()
        es_autismo_paciente = (dx_paciente == "F840")
        es_deglucion_paciente = (cups_paciente == "937203")
        es_individual_paciente = es_autismo_paciente or es_deglucion_paciente
        
        clave_nueva = f"{data['nuevo_profesional']}|{fecha}|{hora}"
        citas_en_slot = agenda_actual.get(clave_nueva, [])
        tiene_autismo_en_slot = any(cita.get("_es_autismo", False) for cita in citas_en_slot)
        tiene_deglucion_en_slot = any(cita.get("_es_deglucion", False) for cita in citas_en_slot)
        tiene_individual_en_slot = tiene_autismo_en_slot or tiene_deglucion_en_slot
        ocupados = len(citas_en_slot)
        
        if es_individual_paciente:
            if ocupados > 0:
                conflictos.append(f"{fecha} {hora} → Paciente requiere sesión individual (Autismo/Deglución). Slot ocupado.")
        else:
            if tiene_individual_en_slot:
                conflictos.append(f"{fecha} {hora} → Slot tiene paciente individual. Sesión exclusiva.")
            elif ocupados >= 2:
                conflictos.append(f"{fecha} {hora} → Horario lleno (máx. 2 pacientes).")
    
    if conflictos:
        mensaje = "❌ **Reasignación BLOQUEADA** por violación de reglas:\n" + "\n".join(conflictos[:10])
        return jsonify({"exito": False, "mensaje": mensaje})
    
    # ✅ SEGUNDA LECTURA: Reconstruir archivo COMPLETO
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        
        for row in reader:
            if (row.get("Tipo_Doc") == data["tipo_doc"] and
                row.get("Num_Doc") == data["num_doc"] and
                row.get("CUPS") == data["cups"] and
                row.get("Profesional") == data["profesional_origen"] and
                row.get("Estado", "activa") == "activa"):
                
                fecha_cita = row.get("Fecha", "")
                
                # ✅ FILTRO CRÍTICO: ¿La cita es pasada?
                if fecha_cita < fecha_hoy:
                    citas_pasadas_preservadas += 1
                    registros_actualizados.append(row)
                    continue
                
                # ✅ Solo reasignar citas futuras
                row["Profesional"] = data["nuevo_profesional"]
                citas_reasignadas += 1
            
            # ✅ CRÍTICO: TODAS las filas se agregan
            registros_actualizados.append(row)
    
    with open("data/agenda.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(registros_actualizados)
    
    # ✅ MENSAJE MEJORADO
    mensaje = f"✅ {citas_reasignadas} citas FUTURAS reasignadas a '{data['nuevo_profesional']}'."
    if citas_pasadas_preservadas > 0:
        mensaje += f" 📁 {citas_pasadas_preservadas} citas PASADAS preservadas con '{data['profesional_origen']}'."
    
    return jsonify({"exito": True, "mensaje": mensaje})
@app.route("/cancelar-bloque", methods=["POST"])
@login_requerido
def cancelar_bloque():
    data = request.get_json()
    required = ["tipo_doc", "num_doc", "profesional", "cups"]
    if not all(k in data for k in required):
        return jsonify({"exito": False, "mensaje": "Faltan datos del bloque."})
    if not os.path.exists("data/agenda.csv"):
        return jsonify({"exito": False, "mensaje": "No existe agenda registrada."})
    registros_actualizados = []
    bloque_cancelado = False
    fieldnames = None
    with open("data/agenda.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or CAMPOS_AGENDA
        for row in reader:
            if (row.get("Tipo_Doc") == data["tipo_doc"] and
                row.get("Num_Doc") == data["num_doc"] and
                row.get("Profesional") == data["profesional"] and
                row.get("CUPS") == data["cups"] and
                row.get("Estado", "activa") == "activa"):
                row["Estado"] = "cancelada"
                bloque_cancelado = True
            elif "Estado" not in row:
                row["Estado"] = "activa"
            registros_actualizados.append(row)
    if not bloque_cancelado:
        return jsonify({"exito": False, "mensaje": "No se encontró un bloque activo con esos datos."})
    with open("data/agenda.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(registros_actualizados)
    return jsonify({"exito": True, "mensaje": "✅ Bloque terapéutico cancelado correctamente."})
# ✅ CORREGIDO: Validación de paciente duplicado en concentradas (MODIFICADA - Deglución = Individual como Autismo)
@app.route("/guardar-citas-concentradas", methods=["POST"])
@login_requerido
def guardar_citas_concentradas():
    data = request.get_json()
    required = ["paciente", "dx", "celular", "email", "servicios", "profesional", "fecha", "hora_inicio"]
    if not all(k in data for k in required):
        return jsonify({"exito": False, "mensaje": "Faltan datos."})
    paciente = data["paciente"]
    dx_paciente = data["dx"]["dx_codigo"].upper()
    cups_paciente = data["servicios"][0].get("cups", "") if data["servicios"] else ""
    es_autismo = dx_paciente == "F840"
    es_deglucion = cups_paciente == "937203"  # ✅ DEGLUCIÓN = INDIVIDUAL
    es_individual = es_autismo or es_deglucion  # ✅ AMBOS REQUIEREN INDIVIDUAL
    profesional = data["profesional"]
    fecha = data["fecha"]
    hora_inicio = data["hora_inicio"]
    servicios = data["servicios"]
    if not es_dia_habil(fecha):
        return jsonify({"exito": False, "mensaje": "La fecha seleccionada no es hábil."})
    total_bloques = len(servicios)
    if hora_inicio not in TODOS_BLOQUES:
        return jsonify({"exito": False, "mensaje": "Hora de inicio inválida."})
    idx_inicio = TODOS_BLOQUES.index(hora_inicio)
    if idx_inicio + total_bloques > len(TODOS_BLOQUES):
        return jsonify({"exito": False, "mensaje": "No hay suficientes bloques horarios consecutivos en el día."})
    horas_necesarias = TODOS_BLOQUES[idx_inicio : idx_inicio + total_bloques]
    agenda_actual = cargar_agenda_desde_csv_filtrada()
    conflictos = []
    for i, hora in enumerate(horas_necesarias):
        clave = f"{profesional}|{fecha}|{hora}"
        citas_existentes = agenda_actual.get(clave, [])
        tiene_autismo_existente = any(cita.get("_es_autismo") for cita in citas_existentes)
        tiene_deglucion_existente = any(cita.get("_es_deglucion") for cita in citas_existentes)  # ✅ AGREGADO
        tiene_individual_existente = tiene_autismo_existente or tiene_deglucion_existente  # ✅ VERIFICAR AMBOS
        ocupados = len(citas_existentes)
        # ✅ VALIDACIÓN DE PACIENTE DUPLICADO
        if os.path.exists("data/agenda.csv"):
            with open("data/agenda.csv", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if (row.get("Num_Doc") == paciente["num_doc"] and
                        row.get("Tipo_Doc") == paciente["tipo_doc"] and
                        row.get("Fecha") == fecha and
                        row.get("Hora") == hora and
                        row.get("Estado", "activa") == "activa" and
                        row.get("Profesional") != profesional):
                        conflictos.append(f"{hora} → Paciente ya agendado con '{row.get('Profesional')}'")
                        break
        if es_individual:  # ✅ AMBOS (autismo y deglución) requieren individual
            if ocupados > 0:
                conflictos.append(f"{hora} → Ya hay cita(s). Paciente requiere sesión individual (Autismo/Deglución).")
        else:
            if tiene_individual_existente:  # ✅ VERIFICAR AMBOS
                conflictos.append(f"{hora} → Ya hay paciente que requiere sesión individual (Autismo/Deglución). Sesión exclusiva.")
            elif ocupados >= 2:
                conflictos.append(f"{hora} → Horario lleno (máx. 2 pacientes).")
    if conflictos:
        mensaje = "❌ No se puede agendar la agenda concentrada:\n" + "\n".join(conflictos)
        return jsonify({"exito": False, "mensaje": mensaje})
    with open("data/agenda.csv", "a", newline="", encoding="utf-8") as f:
        writer = None
        for i, hora in enumerate(horas_necesarias):
            servicio = servicios[i]
            cita = {
                "Tipo_Doc": paciente["tipo_doc"],
                "Num_Doc": paciente["num_doc"],
                "Nombre_Completo": paciente["nombre_completo"],
                "Edad": str(paciente.get("edad") or ""),
                "Genero": paciente["genero"],
                "REGIMEN": paciente.get("REGIMEN", ""),
                "TIPO_AFILIADO": paciente.get("TIPO_AFILIADO", ""),
                "CODIGO_EPS": paciente.get("CODIGO_EPS", ""),
                "DEPARTAMENTO": paciente.get("DEPARTAMENTO", ""),
                "CIUDAD": paciente.get("CIUDAD", ""),
                "ZONA": paciente.get("ZONA", ""),
                "ESTADO": paciente.get("ESTADO", ""),
                "Celular": data["celular"],
                "Email": data["email"],
                "Dx_Codigo": data["dx"]["dx_codigo"],
                "Dx_Descripcion": data["dx"]["descripcion"],
                "CUPS": servicio["cups"],
                "Servicio": servicio["nombre"],
                "Profesional": profesional,
                "Fecha": fecha,
                "Hora": hora,
                "Cantidad_Total": str(total_bloques),
                "Frecuencia_Semanal": "0",
                "Duracion_Meses": "0",
                "Observacion": "Agenda concentrada - múltiples servicios",
                "Fecha_Registro": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "Estado": "activa"
            }
            cita_completa = {campo: cita.get(campo, "") for campo in CAMPOS_AGENDA}
            if writer is None:
                existe = os.path.exists("data/agenda.csv") and os.path.getsize("data/agenda.csv") > 0
                writer = csv.DictWriter(f, fieldnames=CAMPOS_AGENDA)
                if not existe:
                    writer.writeheader()
            writer.writerow(cita_completa)
    return jsonify({"exito": True, "mensaje": f"✅ Agenda concentrada guardada: {total_bloques} sesiones asignadas consecutivamente."})
@app.route("/buscar-horario-viable", methods=["POST"])
@login_requerido
def buscar_horario_viable():
    data = request.get_json()
    required = ["paciente", "dx", "servicio", "profesional", "fecha_inicio"]
    if not all(k in data for k in required):
        return jsonify({"exito": False, "mensaje": "Faltan datos."})
    paciente = data["paciente"]
    servicio = data["servicio"]
    profesional = data["profesional"]
    fecha_inicio = data["fecha_inicio"]
    cantidad_total = int(servicio.get("cantidad_total", 1))
    frecuencia_semanal = int(servicio.get("frecuencia_semanal", 1))
    dx_paciente = data["dx"]["dx_codigo"].upper()
    cups_paciente = servicio.get("cups", "")
    es_autismo = dx_paciente == "F840"
    es_deglucion = cups_paciente == "937203"  # ✅ DEGLUCIÓN = INDIVIDUAL
    es_individual = es_autismo or es_deglucion  # ✅ AMBOS REQUIEREN INDIVIDUAL
    fechas = generar_fechas_terapia(fecha_inicio, frecuencia_semanal, cantidad_total)
    fechas = list(dict.fromkeys(fechas))
    if not fechas:
        return jsonify({"exito": False, "mensaje": "No se pudieron generar fechas válidas."})
    fechas_no_habiles = [f for f in fechas if not es_dia_habil(f)]
    if fechas_no_habiles:
        return jsonify({"exito": False, "mensaje": "Fechas no hábiles en el bloque."})
    agenda_actual = cargar_agenda_desde_csv_filtrada()
    for hora in TODOS_BLOQUES:
        viable = True
        for fecha in fechas:
            clave = f"{profesional}|{fecha}|{hora}"
            citas = agenda_actual.get(clave, [])
            ocupados = len(citas)
            tiene_autismo_existente = any(cita.get("_es_autismo") for cita in citas)
            tiene_deglucion_existente = any(cita.get("_es_deglucion") for cita in citas)  # ✅ AGREGADO
            tiene_individual_existente = tiene_autismo_existente or tiene_deglucion_existente  # ✅ VERIFICAR AMBOS
            if es_individual:  # ✅ AMBOS (autismo y deglución)
                if ocupados > 0:
                    viable = False
                    break
            else:
                if tiene_individual_existente or ocupados >= 2:  # ✅ VERIFICAR AMBOS
                    viable = False
                    break
        if viable:
            return jsonify({"exito": True, "hora_sugerida": hora, "fechas": fechas, "mensaje": f"✅ Horario viable encontrado: {hora}. ¿Desea agendar todo el bloque aquí?"})
    return jsonify({"exito": False, "mensaje": "❌ No se encontró ningún horario disponible para agendar todo el bloque completo."})
@app.route("/generar-reporte-whatsapp-pdf", methods=["POST"])
@login_requerido
def generar_reporte_whatsapp_pdf():
    data = request.get_json()
    tipo_doc = data.get("tipo_doc", "").strip().upper()
    num_doc = data.get("num_doc", "").strip()
    if not tipo_doc or not num_doc:
        return jsonify({"error": "Documento requerido"}), 400
    servicios = {}
    try:
        if os.path.exists("data/SERVICIOS_ISS.csv"):
            with open("data/SERVICIOS_ISS.csv", encoding="utf-8") as f:
                reader = csv.DictReader(f, delimiter="|")
                for row in reader:
                    servicios[row["CUPS_ISS"].strip()] = row["SERVICIO"].strip()
    except Exception as e:
        print("Error al cargar servicios:", e)
    hoy = date.today()
    paciente_info = None
    citas_paciente = []
    try:
        with open("data/agenda.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if (row.get("Tipo_Doc") == tipo_doc and
                    row.get("Num_Doc") == num_doc and
                    row.get("Estado", "activa") == "activa"):
                    if paciente_info is None:
                        paciente_info = {
                            "nombre_completo": row.get("Nombre_Completo", ""),
                            "edad": str(row.get("Edad", "") or "?"),
                            "ciudad": row.get("CIUDAD", ""),
                            "documento": num_doc
                        }
                    try:
                        fecha_cita = datetime.strptime(row["Fecha"], "%Y-%m-%d").date()
                        if fecha_cita >= hoy:
                            cups = row.get("CUPS", "").strip()
                            servicio = servicios.get(cups, cups) if cups else "Sin servicio"
                            citas_paciente.append({"fecha": row["Fecha"], "hora": row["Hora"], "profesional": row.get("Profesional", "").strip() or "Sin asignar", "servicio": servicio})
                    except:
                        continue
    except Exception as e:
        return jsonify({"error": f"Error al leer agenda: {str(e)}"}), 500
    if not paciente_info:
        return jsonify({"error": "Paciente no encontrado"}), 404
    citas_paciente.sort(key=lambda x: (x["fecha"], x["hora"]))
    pdf = ReportePDF()
    pdf.add_page()
    lineas_paciente = [
        f"Nombre: {paciente_info['nombre_completo']}",
        f"Documento: {tipo_doc}-{paciente_info['documento']}",
        f"Edad: {paciente_info['edad']}",
        f"Ciudad: {paciente_info['ciudad']}"
    ]
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "DATOS DEL PACIENTE", 0, 1, "L")
    pdf.set_font("Arial", "", 10)
    for linea in lineas_paciente:
        pdf.cell(0, 6, linea, 0, 1, "L")
    pdf.ln(4)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "TERAPIAS ASIGNADAS", 0, 1, "L")
    pdf.set_font("Arial", "", 10)
    if not citas_paciente:
        pdf.cell(0, 6, "No hay terapias programadas futuras.", 0, 1, "L")
    else:
        for cita in citas_paciente:
            fecha_formateada = datetime.strptime(cita["fecha"], "%Y-%m-%d").strftime("%d/%m/%Y")
            linea = f"{fecha_formateada} - {cita['hora']} -> {cita['servicio']} -> {cita['profesional']}"
            pdf.cell(0, 6, linea, 0, 1, "L")
    pdf.ln(6)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 8, "REQUISITOS IMPORTANTES", 0, 1, "L")
    pdf.set_font("Arial", "", 10)
    requisitos = [
        "- Presentar documento original del paciente.",
        "- Llegar 10 minutos antes de la cita.",
        "- Solo se permite un acompañante adulto."
    ]
    for req in requisitos:
        pdf.cell(0, 6, req, 0, 1, "L")
    buffer = BytesIO()
    pdf_output = pdf.output()
    buffer.write(pdf_output)
    buffer.seek(0)
    nombre_archivo = f"reporte_neuro_{tipo_doc}_{num_doc}.pdf"
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=nombre_archivo)

@app.route("/terapeuta/gas")
@login_requerido
def terapeuta_gas_panel():
    """Panel exclusivo para profesionales: ver sus pacientes y gestionar metas GAS"""
    if session.get('rol') not in ['admin', 'profesional']:
        return redirect(url_for('index'))
    _iniciar_csv_gas()
    profesional = session.get('usuario')
    # Pacientes con citas activas asignadas a este profesional
    pacientes = set()
    if os.path.exists("data/agenda.csv"):
        with open("data/agenda.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("Profesional") == profesional and row.get("Estado") == "activa":
                    pacientes.add((row["Tipo_Doc"], row["Num_Doc"], row["Nombre_Completo"]))
    return render_template("terapeuta_gas.html", 
                          pacientes=sorted(pacientes, key=lambda x: x[2]),
                          profesional=profesional)

@app.route("/gas/crear", methods=["POST"])
@login_requerido
def gas_crear_meta():
    """Crear una nueva meta GAS para un paciente (con línea base)"""
    if session.get('rol') != 'profesional':
        return jsonify({"exito": False, "mensaje": "Acceso denegado."})
    _iniciar_csv_gas()
    
    data = request.get_json()
    required = ["tipo_doc", "num_doc", "descripcion", "dominio", "peso", 
                "fecha_evaluacion", "niveles", "linea_base"]  # ✅ NUEVO: linea_base
    if not all(k in data for k in required):
        return jsonify({"exito": False, "mensaje": "Faltan datos obligatorios"})
    
    meta_id = f"GAS{datetime.now().strftime('%Y%m%d%H%M%S')}"
    linea_base = int(data["linea_base"])  # -2, -1, 0, +1, +2
    
    # Guardar meta principal CON línea base
    with open("data/gas_metas.csv", "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="|")
        writer.writerow([
            meta_id, data["tipo_doc"], data["num_doc"], data["descripcion"],
            data["dominio"], data["peso"], datetime.now().strftime("%Y-%m-%d"),
            data["fecha_evaluacion"], "activa", session.get("usuario", ""),
            linea_base  # ✅ GUARDAR LÍNEA BASE
        ])
    
    # Guardar los 5 niveles
    with open("data/gas_niveles.csv", "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="|")
        for nivel in data["niveles"]:
            writer.writerow([meta_id, nivel["puntuacion"], nivel["criterio"], 
                           nivel.get("metodo", "observacion_directa")])
    
    # ✅ REGISTRAR EVALUACIÓN INICIAL AUTOMÁTICAMENTE
    eval_id = f"EVAL{datetime.now().strftime('%Y%m%d%H%M%S')}"
    with open("data/gas_evaluaciones.csv", "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="|")
        writer.writerow([
            eval_id, meta_id, datetime.now().strftime("%Y-%m-%d"),
            linea_base, "Línea base inicial", session.get("usuario", ""), "inicial"
        ])
    
    return jsonify({"exito": True, "mensaje": "Meta GAS creada con línea base", "meta_id": meta_id})
@app.route("/gas/listar/<tipo_doc>/<num_doc>")
@login_requerido
def gas_listar_metas(tipo_doc, num_doc):
    """Listar metas GAS de un paciente"""
    if session.get('rol') != 'profesional':
        return jsonify({"metas": []})
    
    metas = []
    if os.path.exists("data/gas_metas.csv"):
        with open("data/gas_metas.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                if row["Paciente_TipoDoc"] == tipo_doc and row["Paciente_NumDoc"] == num_doc:
                    niveles = []
                    if os.path.exists("data/gas_niveles.csv"):
                        with open("data/gas_niveles.csv", encoding="utf-8") as f2:
                            reader2 = csv.DictReader(f2, delimiter="|")
                            for row2 in reader2:
                                if row2["Meta_ID"] == row["Meta_ID"]:
                                    # ✅ CONVERTIR None a valores por defecto
                                    niveles.append({
                                        "puntuacion": int(row2["Puntuacion"]) if row2["Puntuacion"] else 0,
                                        "criterio": row2["Criterio_Observable"] or ""
                                    })
                    
                    # ✅ SORT SEGURO (evitar None en puntuacion)
                    niveles.sort(key=lambda x: x["puntuacion"] if x["puntuacion"] is not None else 0)
                    
                    # ✅ CONVERTIR TODOS LOS CAMPOS None A STRING VACÍO
                    meta_segura = {
                        "Meta_ID": row["Meta_ID"] or "",
                        "Paciente_TipoDoc": row["Paciente_TipoDoc"] or "",
                        "Paciente_NumDoc": row["Paciente_NumDoc"] or "",
                        "Descripcion": row["Descripcion"] or "",
                        "Dominio": row["Dominio"] or "",
                        "Peso": int(row["Peso"]) if row["Peso"] else 1,
                        "Fecha_Creacion": row["Fecha_Creacion"] or "",
                        "Fecha_Evaluacion": row["Fecha_Evaluacion"] or "",
                        "Estado": row["Estado"] or "activa",
                        "Profesional_Creacion": row["Profesional_Creacion"] or "",
                        "Linea_Base_Puntuacion": int(row["Linea_Base_Puntuacion"]) if row.get("Linea_Base_Puntuacion") else -1,
                        "niveles": niveles
                    }
                    metas.append(meta_segura)
    
    return jsonify({"metas": metas})

@app.route("/gas/evaluar", methods=["POST"])
@login_requerido
def gas_evaluar_meta():
    """Registrar evaluación de una meta GAS"""
    if session.get('rol') != 'profesional':
        return jsonify({"exito": False, "mensaje": "Acceso denegado."})
    data = request.get_json()
    required = ["meta_id", "puntuacion_lograda", "evidencia"]
    if not all(k in data for k in required):
        return jsonify({"exito": False, "mensaje": "Faltan datos"})
    
    eval_id = f"EVAL{datetime.now().strftime('%Y%m%d%H%M%S')}"
    with open("data/gas_evaluaciones.csv", "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="|")
        writer.writerow([eval_id, data["meta_id"], datetime.now().strftime("%Y-%m-%d"), data["puntuacion_lograda"], data["evidencia"], session.get("usuario", "")])
    
    # Actualizar estado de la meta
    if os.path.exists("data/gas_metas.csv"):
        registros = []
        with open("data/gas_metas.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="|")
            fieldnames = reader.fieldnames
            for row in reader:
                if row["Meta_ID"] == data["meta_id"]:
                    row["Estado"] = "evaluada"
                registros.append(row)
        with open("data/gas_metas.csv", "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter="|")
            writer.writeheader()
            writer.writerows(registros)
    
    return jsonify({"exito": True, "mensaje": "Evaluación registrada"})

@app.route("/gas/registrar-evaluacion", methods=["POST"])
@login_requerido
def gas_registrar_evaluacion():
    """Registrar evaluación intermedia o final de una meta GAS"""
    if session.get('rol') != 'profesional':
        return jsonify({"exito": False, "mensaje": "Acceso denegado."})
    
    data = request.get_json()
    required = ["meta_id", "puntuacion", "evidencia", "tipo_eval"]
    if not all(k in data for k in required):
        return jsonify({"exito": False, "mensaje": "Faltan datos"})
    
    eval_id = f"EVAL{datetime.now().strftime('%Y%m%d%H%M%S')}"
    with open("data/gas_evaluaciones.csv", "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="|")
        writer.writerow([
            eval_id, data["meta_id"], datetime.now().strftime("%Y-%m-%d"),
            int(data["puntuacion"]), data["evidencia"], 
            session.get("usuario", ""), data["tipo_eval"]
        ])
    
    # Si es evaluación final, actualizar estado de la meta
    if data["tipo_eval"] == "final":
        if os.path.exists("data/gas_metas.csv"):
            registros = []
            with open("data/gas_metas.csv", encoding="utf-8") as f:
                reader = csv.DictReader(f, delimiter="|")
                fieldnames = reader.fieldnames
                for row in reader:
                    if row["Meta_ID"] == data["meta_id"]:
                        row["Estado"] = "finalizada"
                    registros.append(row)
            with open("data/gas_metas.csv", "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter="|")
                writer.writeheader()
                writer.writerows(registros)
    
    return jsonify({"exito": True, "mensaje": f"Evaluación {data['tipo_eval']} registrada"})

@app.route("/gas/historial/<tipo_doc>/<num_doc>")
@login_requerido
def gas_historial_progreso(tipo_doc, num_doc):
    """Obtener historial de evaluaciones para gráfico de progreso"""
    if session.get('rol') != 'profesional':
        return jsonify({"progreso": []})
    
    progreso = []
    
    # Obtener metas del paciente
    metas_ids = []
    if os.path.exists("data/gas_metas.csv"):
        with open("data/gas_metas.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                if row["Paciente_TipoDoc"] == tipo_doc and row["Paciente_NumDoc"] == num_doc:
                    metas_ids.append(row["Meta_ID"])
    
    # Obtener evaluaciones de esas metas
    if os.path.exists("data/gas_evaluaciones.csv") and metas_ids:
        with open("data/gas_evaluaciones.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                if row["Meta_ID"] in metas_ids:
                    progreso.append({
                        "meta_id": row["Meta_ID"],
                        "fecha": row["Fecha_Eval"],
                        "puntuacion": int(row["Puntuacion_Lograda"]),
                        "tipo": row["Tipo_Eval"],
                        "evidencia": row["Evidencia"]
                    })
    
    # Ordenar por fecha
    progreso.sort(key=lambda x: x["fecha"])
    
    return jsonify({"progreso": progreso})
@app.route("/gas/tscore/<tipo_doc>/<num_doc>")
@login_requerido
def gas_obtener_tscore(tipo_doc, num_doc):
    """Obtener T-score GAS consolidado para un paciente"""
    if session.get('rol') != 'profesional':
        return jsonify({"error": "Acceso denegado"})
    
    metas_evaluadas = []
    
    # Cargar evaluaciones
    evaluaciones = {}
    if os.path.exists("data/gas_evaluaciones.csv"):
        with open("data/gas_evaluaciones.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                evaluaciones[row["Meta_ID"]] = {
                    "puntuacion": int(row["Puntuacion_Lograda"]),
                    "evidencia": row["Evidencia"],
                    "fecha": row["Fecha_Eval"]
                }
    
    # Cargar metas del paciente con sus pesos
    if os.path.exists("data/gas_metas.csv"):
        with open("data/gas_metas.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                if (row["Paciente_TipoDoc"] == tipo_doc and 
                    row["Paciente_NumDoc"] == num_doc and
                    row["Meta_ID"] in evaluaciones):
                    metas_evaluadas.append({
                        "meta_id": row["Meta_ID"],
                        "descripcion": row["Descripcion"],
                        "peso": int(row["Peso"]),
                        "puntuacion": evaluaciones[row["Meta_ID"]]["puntuacion"],
                        "evidencia": evaluaciones[row["Meta_ID"]]["evidencia"],
                        "fecha_eval": evaluaciones[row["Meta_ID"]]["fecha"]
                    })
    
    # Calcular T-score
    resultado = calcular_tscore_gas_multiple(metas_evaluadas)
    
    return jsonify({
        "paciente": f"{tipo_doc}-{num_doc}",
        "metas_evaluadas": len(metas_evaluadas),
        "t_score": resultado["t_score"],
        "interpretacion": resultado["interpretacion"],
        "detalles": resultado["detalles"],
        "metas_detalle": metas_evaluadas  # Para mostrar en frontend
    })
# === REPORTE PROFESIONAL GAS ===
@app.route("/gas/reporte/<tipo_doc>/<num_doc>")
@login_requerido
def gas_reporte_paciente(tipo_doc, num_doc):
    """Generar reporte profesional GAS para un paciente"""
    if session.get('rol') not in ['admin', 'profesional']:
        return redirect(url_for('index'))
    
    # Cargar datos del paciente
    paciente_info = {"tipo_doc": tipo_doc, "num_doc": num_doc, "nombre": ""}
    metas = []
    evaluaciones = []
    
    # Buscar nombre en agenda
    if os.path.exists("data/agenda.csv"):
        with open("data/agenda.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row["Tipo_Doc"] == tipo_doc and row["Num_Doc"] == num_doc:
                    paciente_info["nombre"] = row["Nombre_Completo"]
                    break
    
    # Cargar metas DEL PACIENTE
    metas_ids_paciente = []  # ✅ NUEVO: Guardar IDs de metas de este paciente
    if os.path.exists("data/gas_metas.csv"):
        with open("data/gas_metas.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                if row["Paciente_TipoDoc"] == tipo_doc and row["Paciente_NumDoc"] == num_doc:
                    metas.append(row)
                    metas_ids_paciente.append(row["Meta_ID"])  # ✅ Guardar ID
    
    # Cargar evaluaciones SOLO de las metas de este paciente
    if os.path.exists("data/gas_evaluaciones.csv") and metas_ids_paciente:
        with open("data/gas_evaluaciones.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                if row["Meta_ID"] in metas_ids_paciente:  # ✅ FILTRO CORRECTO
                    evaluaciones.append(row)
    
    # Calcular T-score
    metas_evaluadas = []
    evaluaciones_dict = {e["Meta_ID"]: e for e in evaluaciones if e.get("Tipo_Eval") == "final"}
    for meta in metas:
        if meta["Meta_ID"] in evaluaciones_dict:
            metas_evaluadas.append({
                "peso": int(meta["Peso"]),
                "puntuacion": int(evaluaciones_dict[meta["Meta_ID"]]["Puntuacion_Lograda"])
            })
    
    tscore_result = calcular_tscore_gas_multiple(metas_evaluadas) if metas_evaluadas else {"t_score": None, "interpretacion": "Sin evaluaciones finales"}
    
    return render_template("gas_reporte.html",
                         paciente=paciente_info,
                         metas=metas,
                         evaluaciones=evaluaciones,
                         tscore=tscore_result,
                         fecha_reporte=datetime.now().strftime("%Y-%m-%d"))
@app.route("/gas/exportar-excel/<tipo_doc>/<num_doc>")
@login_requerido
def gas_exportar_excel(tipo_doc, num_doc):
    """Exportar reporte GAS a Excel profesional"""
    if session.get('rol') not in ['admin', 'profesional']:
        return redirect(url_for('index'))
    
    # Cargar datos
    metas = []
    evaluaciones = []
    metas_ids_paciente = []  # ✅ NUEVO
    
    if os.path.exists("data/gas_metas.csv"):
        with open("data/gas_metas.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                if row["Paciente_TipoDoc"] == tipo_doc and row["Paciente_NumDoc"] == num_doc:
                    metas.append(row)
                    metas_ids_paciente.append(row["Meta_ID"])  # ✅ Guardar ID
    
    if os.path.exists("data/gas_evaluaciones.csv"):
        with open("data/gas_evaluaciones.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                if row["Meta_ID"] in metas_ids_paciente:  # ✅ FILTRO CORRECTO
                    evaluaciones.append(row)
    
    # Calcular T-score
    metas_evaluadas = []
    evaluaciones_dict = {e["Meta_ID"]: e for e in evaluaciones if e["Tipo_Eval"] == "final"}
    for meta in metas:
        if meta["Meta_ID"] in evaluaciones_dict:
            metas_evaluadas.append({
                "peso": int(meta["Peso"]),
                "puntuacion": int(evaluaciones_dict[meta["Meta_ID"]]["Puntuacion_Lograda"])
            })
    
    tscore_result = calcular_tscore_gas_multiple(metas_evaluadas) if metas_evaluadas else {"t_score": None}
    
    # Crear Excel con múltiples hojas
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Hoja 1: Resumen
        df_resumen = pd.DataFrame([{
            'Paciente': tipo_doc + '-' + num_doc,
            'Fecha Reporte': datetime.now().strftime("%Y-%m-%d"),
            'Total Metas': len(metas),
            'Metas Evaluadas': len(metas_evaluadas),
            'T-Score': tscore_result.get('t_score', 'N/A'),
            'Interpretación': tscore_result.get('interpretacion', 'N/A')
        }])
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
        
        # Hoja 2: Metas
        if metas:
            df_metas = pd.DataFrame(metas)
            df_metas.to_excel(writer, sheet_name='Metas', index=False)
        
        # Hoja 3: Evaluaciones
        if evaluaciones:
            df_evals = pd.DataFrame(evaluaciones)
            df_evals.to_excel(writer, sheet_name='Evaluaciones', index=False)
        
        # Formato profesional
        workbook = writer.book
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            # Auto-ajustar columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    nombre_archivo = f"GAS_Reporte_{tipo_doc}_{num_doc}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre_archivo
    )

# === NUEVA RUTA: Listar pacientes con escala GAS ===
@app.route("/gas/pacientes-con-gas")
@login_requerido
def gas_listar_pacientes():
    """Retorna lista de pacientes únicos que tienen metas GAS registradas"""
    if session.get('rol') not in ['admin', 'profesional']:
        return redirect(url_for('index'))
    
    # ✅ PASO 1: Obtener documentos únicos de pacientes con GAS
    pacientes_docs = set()
    if os.path.exists("data/gas_metas.csv"):
        with open("data/gas_metas.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                clave = f"{row['Paciente_TipoDoc']}|{row['Paciente_NumDoc']}"
                pacientes_docs.add(clave)
    
    # ✅ PASO 2: Buscar nombres en agenda.csv
    nombres_pacientes = {}
    if os.path.exists("data/agenda.csv"):
        with open("data/agenda.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                clave = f"{row['Tipo_Doc']}|{row['Num_Doc']}"
                if clave not in nombres_pacientes and row.get('Nombre_Completo'):
                    nombres_pacientes[clave] = row['Nombre_Completo']
    
    # ✅ PASO 3: Construir lista con nombres
    lista = []
    for doc_clave in pacientes_docs:
        partes = doc_clave.split("|")
        tipo_doc = partes[0]
        num_doc = partes[1]
        nombre = nombres_pacientes.get(doc_clave, "Sin nombre registrado")
        
        lista.append({
            "tipo_doc": tipo_doc,
            "num_doc": num_doc,
            "nombre": nombre
        })
    
    # ✅ Ordenar por nombre
    lista.sort(key=lambda x: x["nombre"])
    
    return jsonify({"exito": True, "pacientes": lista, "total": len(lista)})
@app.route("/gas/reporte-imprimir/<tipo_doc>/<num_doc>")
@login_requerido
def gas_reporte_imprimir(tipo_doc, num_doc):
    """Generar reporte GAS profesional para impresión/PDF"""
    if session.get('rol') not in ['admin', 'profesional']:
        return redirect(url_for('index'))
    
    # Cargar datos del paciente
    paciente_info = {"tipo_doc": tipo_doc, "num_doc": num_doc, "nombre": ""}
    metas = []
    evaluaciones = []
    
    # Buscar nombre en agenda
    if os.path.exists("data/agenda.csv"):
        with open("data/agenda.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row["Tipo_Doc"] == tipo_doc and row["Num_Doc"] == num_doc:
                    paciente_info["nombre"] = row["Nombre_Completo"]
                    break
    
    # Cargar metas DEL PACIENTE
    metas_ids_paciente = []
    if os.path.exists("data/gas_metas.csv"):
        with open("data/gas_metas.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                if row["Paciente_TipoDoc"] == tipo_doc and row["Paciente_NumDoc"] == num_doc:
                    metas.append(row)
                    metas_ids_paciente.append(row["Meta_ID"])
    
    # Cargar evaluaciones SOLO de las metas de este paciente
    if os.path.exists("data/gas_evaluaciones.csv") and metas_ids_paciente:
        with open("data/gas_evaluaciones.csv", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                if row["Meta_ID"] in metas_ids_paciente:
                    evaluaciones.append(row)
    
    # Calcular T-score
    metas_evaluadas = []
    evaluaciones_dict = {e["Meta_ID"]: e for e in evaluaciones if e.get("Tipo_Eval") == "final"}
    for meta in metas:
        if meta["Meta_ID"] in evaluaciones_dict:
            metas_evaluadas.append({
                "peso": int(meta["Peso"]),
                "puntuacion": int(evaluaciones_dict[meta["Meta_ID"]]["Puntuacion_Lograda"])
            })
    
    tscore_result = calcular_tscore_gas_multiple(metas_evaluadas) if metas_evaluadas else {"t_score": None, "interpretacion": "Sin evaluaciones finales"}
    
    return render_template("gas_reporte_imprimir.html",
                         paciente=paciente_info,
                         metas=metas,
                         evaluaciones=evaluaciones,
                         tscore=tscore_result,
                         fecha_reporte=datetime.now().strftime("%Y-%m-%d"),
                         profesional=session.get('usuario', ''))

if __name__ == "__main__":
    print("✅ ================================================")
    print("✅ Agendamiento-Med Elite v2.0 - Acceso por Profesional")
    print("✅ Sistema listo para red local y celulares")
    print("✅ Profesionales: Login por nombre (sin contraseña)")
    print("✅ Admin: Login con credenciales")
    print("✅ ================================================")
    print("📱 Para acceder desde celulares:")
    print("   1. Conéctense al mismo WiFi de la clínica")
    print("   2. Abran navegador y escriban: http://[IP_DEL_SERVIDOR]:5000")
    print("   3. Seleccionen su nombre y ¡listo!")
    print("✅ ================================================")
    app.run(debug=False, host="0.0.0.0", port=5000)